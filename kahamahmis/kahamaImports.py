import openpyxl
from django.shortcuts import render
from django.http import HttpResponseRedirect
from django.urls import reverse
from django.contrib import messages
from clinic.models import Country, DiseaseRecode, HealthRecord, InsuranceCompany, PathodologyRecord, RemoteCompany, RemoteMedicine, RemoteService
from .forms import CountryImportForm, DiseaseRecodeImportForm, HealthRecordImportForm, ImportInsuranceCompanyForm, PathodologyRecordImportForm, RemoteCompanyImportForm, RemoteMedicineImportForm, RemoteServiceImportForm
from django.db.utils import IntegrityError

def import_insurance_company_data(request):
    if request.method == 'POST':
        form = ImportInsuranceCompanyForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            # Read headers
            headers = [cell.value for cell in sheet[1]]
            required_headers = ['name', 'phone', 'short_name', 'email', 'address', 'website']

            # Validate headers
            if headers[:len(required_headers)] != required_headers:
                messages.error(request, 'Invalid file format')
                return render(request, 'kahamaImport/import_insurance_company.html', {'form': form})

            # Read data from rows
            for row in sheet.iter_rows(min_row=2, values_only=True):
                data = dict(zip(headers, row))
                try:
                    InsuranceCompany.objects.create(
                        name=data['name'],
                        phone=data['phone'],
                        short_name=data['short_name'],
                        email=data['email'],
                        address=data['address'],
                        website=data['website']
                    )
                except IntegrityError:
                    # Skip duplicate entries and continue
                    continue

            return HttpResponseRedirect(reverse('kahamahmis:manage_insurance'))

    else:
        form = ImportInsuranceCompanyForm()

    return render(request, 'kahamaImport/import_insurance_company.html', {'form': form})

def import_disease_recode_data(request):
    if request.method == 'POST':
        form = DiseaseRecodeImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            # Read headers
            headers = [cell.value for cell in sheet[1]]
            required_headers = ['disease_name', 'code']

            # Validate headers
            if headers[:len(required_headers)] != required_headers:
                messages.error(request, 'Invalid file format')
                return render(request, 'kahamaImport/import_disease_recode.html', {'form': form})

            # Read data from rows
            for row in sheet.iter_rows(min_row=2, values_only=True):
                data = dict(zip(headers, row))
                try:
                    DiseaseRecode.objects.create(
                        disease_name=data['disease_name'],
                        code=data['code']
                    )
                except IntegrityError:
                    # Skip duplicate entries and continue
                    continue

            return HttpResponseRedirect(reverse('kahamahmis:manage_disease'))

    else:
        form = DiseaseRecodeImportForm()

    return render(request, 'kahamaImport/import_disease_recode.html', {'form': form})


def import_remote_medicine_data(request):
    if request.method == 'POST':
        form = RemoteMedicineImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            # Read headers
            headers = [cell.value for cell in sheet[1]]
            required_headers = [
                'drug_name', 'drug_type', 'formulation_unit', 'manufacturer',
                'quantity', 'dividable', 'batch_number', 'expiration_date', 
                'unit_cost', 'buying_price'
            ]

            # Validate headers
            if headers[:len(required_headers)] != required_headers:
                messages.error(request, 'Invalid file format')
                return render(request, 'kahamaImport/import_remote_medicine.html', {'form': form})

            # Read data from rows
            for row in sheet.iter_rows(min_row=2, values_only=True):
                data = dict(zip(headers, row))
                try:
                    RemoteMedicine.objects.create(
                        drug_name=data['drug_name'],
                        drug_type=data['drug_type'],
                        formulation_unit=data['formulation_unit'],
                        manufacturer=data['manufacturer'],
                        quantity=data['quantity'],
                        dividable=data['dividable'],
                        batch_number=data['batch_number'],
                        expiration_date=data['expiration_date'],
                        unit_cost=data['unit_cost'],
                        buying_price=data['buying_price'],
                        total_buying_price=float(data['buying_price']) * data['quantity'] if data['buying_price'] and data['quantity'] else None,
                        remain_quantity=data['quantity'],
                    )
                except IntegrityError:
                    # Skip duplicate entries and continue
                    continue

            return HttpResponseRedirect(reverse('kahamahmis:remotemedicine_list'))

    else:
        form = RemoteMedicineImportForm()

    return render(request, 'kahamaImport/import_remote_medicine.html', {'form': form})


def import_health_record_data(request):
    if request.method == 'POST':
        form = HealthRecordImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            # Read headers
            headers = [cell.value for cell in sheet[1]]
            required_headers = ['name']

            # Validate headers
            if headers[:len(required_headers)] != required_headers:
                messages.error(request, 'Invalid file format')
                return render(request, 'kahamaImport/import_health_record.html', {'form': form})

            # Read data from rows
            for row in sheet.iter_rows(min_row=2, values_only=True):
                data = dict(zip(headers, row))
                # Use get_or_create to avoid duplicates
                HealthRecord.objects.get_or_create(
                    name=data['name']
                )

            return HttpResponseRedirect(reverse('kahamahmis:health_record_list'))

    else:
        form = HealthRecordImportForm()

    return render(request, 'kahamaImport/import_health_record.html', {'form': form})


def import_remote_company_data(request):
    if request.method == 'POST':
        form = RemoteCompanyImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            # Read headers
            headers = [cell.value for cell in sheet[1]]
            required_headers = ['name', 'industry', 'sector', 'headquarters', 'Founded', 'Notes']

            # Validate headers
            if headers[:len(required_headers)] != required_headers:
                messages.error(request, 'Invalid file format')
                return render(request, 'kahamaImport/import_remote_company.html', {'form': form})

            # Read data from rows
            for row in sheet.iter_rows(min_row=2, values_only=True):
                data = dict(zip(headers, row))
                
                # Check if the required fields are present and default missing optional fields to empty string
                name = data.get('name')
                industry = data.get('industry', '')
                sector = data.get('sector', '')
                headquarters = data.get('headquarters', '')
                Founded = data.get('Founded', '')
                Notes = data.get('Notes', '')

                if name:  # Ensure that the name field is present before creating the record
                    try:
                        RemoteCompany.objects.create(
                            name=name,
                            industry=industry,
                            sector=sector,
                            headquarters=headquarters,
                            Founded=Founded,
                            Notes=Notes
                        )
                    except IntegrityError:
                        # Skip duplicate entries and continue
                        continue

            return HttpResponseRedirect(reverse('kahamahmis:manage_company'))
    else:
        form = RemoteCompanyImportForm()

    return render(request, 'kahamaImport/import_remote_company.html', {'form': form})



def import_pathodology_record_data(request):
    if request.method == 'POST':
        form = PathodologyRecordImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            # Read headers
            headers = [cell.value for cell in sheet[1]]
            required_headers = ['name', 'description']

            # Validate headers
            if headers[:len(required_headers)] != required_headers:
                messages.error(request, 'Invalid file format')
                return render(request, 'kahamaImport/import_pathodology_record.html', {'form': form})

            # Read data from rows
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Ensure each header is mapped to a value, defaulting to an empty string if necessary
                data = dict(zip(headers, row))
                name = data.get('name')
                description = data.get('Description', '')

                if name:  # Only proceed if name is present
                    try:
                        PathodologyRecord.objects.create(
                            name=name,
                            description=description
                        )
                    except IntegrityError:
                        # Skip duplicate entries and continue
                        continue

            return HttpResponseRedirect(reverse('kahamahmis:manage_pathodology'))
    else:
        form = PathodologyRecordImportForm()
    return render(request, 'kahamaImport/import_pathodology_record.html', {'form': form})

def import_remote_service_data(request):
    if request.method == 'POST':
        form = RemoteServiceImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            # Read headers
            headers = [cell.value for cell in sheet[1]]
            required_headers = ['name', 'description', 'category']

            # Validate headers
            if headers[:len(required_headers)] != required_headers:
                messages.error(request, 'Invalid file format')
                return render(request, 'kahamaImport/import_remote_service.html', {'form': form})

            # Read data from rows
            for row in sheet.iter_rows(min_row=2, values_only=True):
                data = dict(zip(headers, row))
                try:
                    RemoteService.objects.create(
                        name=data['name'],
                        description=data['description'],
                        category=data['category']
                    )
                except IntegrityError:
                    # Skip duplicate entries and continue
                    continue

            return HttpResponseRedirect(reverse('kahamahmis:remoteservice_list'))

    else:
        form = RemoteServiceImportForm()

    return render(request, 'kahamaImport/import_remote_service.html', {'form': form})

def import_country_data(request):
    if request.method == 'POST':
        form = CountryImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            # Read headers
            headers = [cell.value for cell in sheet[1]]
            required_headers = ['name']

            # Validate headers
            if headers[:len(required_headers)] != required_headers:
                messages.error(request, 'Invalid file format')
                return render(request, 'kahamaImport/import_country.html', {'form': form})

            # Read data from rows
            for row in sheet.iter_rows(min_row=2, values_only=True):
                data = dict(zip(headers, row))
                try:
                    Country.objects.create(
                        name=data['name']
                    )
                except IntegrityError:
                    # Skip duplicate entries and continue
                    continue

            return HttpResponseRedirect(reverse('kahamahmis:manage_country'))

    else:
        form = CountryImportForm()

    return render(request, 'kahamaImport/import_country.html', {'form': form})