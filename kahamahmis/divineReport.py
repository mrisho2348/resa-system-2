from django.shortcuts import render
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from openpyxl import Workbook
from django.db.models.functions import ExtractMonth
from django.db.models import Count
from datetime import datetime
from clinic.forms import YearMonthSelectionForm
from clinic.models import PathodologyRecord, RemoteCompany, RemoteLaboratoryOrder, RemotePatient, RemoteProcedure, RemoteReferral, RemoteService
from openpyxl.utils import get_column_letter
from django.db.models.functions import TruncDay
from calendar import  monthrange
import calendar
import traceback
from django.contrib import messages

def fetch_pathology_reports(year):
    # Get all distinct Pathology Record names for the given year
    all_pathology_records = PathodologyRecord.objects.values_list('name', flat=True)

    # Query the database to get patient counts grouped by Pathology Record and month for the given year
    patients_by_pathology_record = (
        PathodologyRecord.objects.filter(remoteconsultationnotes__created_at__year=year)
        .annotate(month=ExtractMonth('remoteconsultationnotes__created_at'))
        .values('name', 'month')
        .annotate(num_patients=Count('remoteconsultationnotes__id'))
    )

    # Organize the data into a dictionary
    pathology_record_reports = {record: [0] * 13 for record in all_pathology_records}  # Add one more column for the total
    month_totals = [0] * 12

    for patient in patients_by_pathology_record:
        pathology_record_name = patient['name']
        month = patient['month']
        num_patients = patient['num_patients']

        if month is not None:
            month_index = month - 1  # ExtractMonth returns month as an integer
            pathology_record_reports[pathology_record_name][month_index] = num_patients
            pathology_record_reports[pathology_record_name][-1] += num_patients  # Add to the total column
            month_totals[month_index] += num_patients  # Add to the total row

    return pathology_record_reports, month_totals

def render_pathology_report(sheet, year):
    # Fetch pathology report data for the given year
    pathology_reports, month_totals = fetch_pathology_reports(year)

    # Add title
    title_font = Font(size=14, bold=True, color="000000")
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = f'Site Name:'
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')
    # Add subtitle
    subtitle_font = Font(size=12, italic=True, color="808080")
    subtitle_cell = sheet.cell(row=2, column=1)
    subtitle_cell.value = 'Patient Counts by Pathology Record and Month'
    subtitle_cell.font = subtitle_font
    subtitle_cell.alignment = Alignment(horizontal='center')

    # Merge subtitle cells
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=14)  

    # Add blank row
    blank_row_cell = sheet.cell(row=3, column=1)
    blank_row_cell.value = ''
    blank_row_cell.alignment = Alignment(horizontal='center')

    # Add headers
    headers = ['CONSULTATIONS / PATHOLOGY'] + [datetime.strptime(str(month), '%m').strftime('%B') for month in range(1, 13)] + ['Total']  
    for col, header in enumerate(headers, start=1):
        header_cell = sheet.cell(row=4, column=col, value=header)
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center')

    # Add data
    row = 5
    for pathology_name, counts_by_month in pathology_reports.items():
        record_cell = sheet.cell(row=row, column=1, value=pathology_name)
        record_cell.alignment = Alignment(horizontal='left')  # Adjust alignment to left

        for col, count in enumerate(counts_by_month, start=2):
            count_cell = sheet.cell(row=row, column=col, value=count)
            count_cell.alignment = Alignment(horizontal='center')

        row += 1

    # Add total row
    total_row = row  # Store the current row number for total row
    total_row_cell = sheet.cell(row=row, column=1, value='Total')
    total_row_cell.font = Font(bold=True)
    total_row_cell.alignment = Alignment(horizontal='center')
    
    for col, total in enumerate(month_totals, start=2):
        total_cell = sheet.cell(row=row, column=col, value=total)
        total_cell.alignment = Alignment(horizontal='center')
    
    # Add total column    
    total_column_cell = sheet.cell(row=4, column=len(headers), value='Total')
    total_column_cell.font = Font(bold=True)
    total_column_cell.alignment = Alignment(horizontal='center')
    for row, (pathology_name, counts) in enumerate(pathology_reports.items(), start=5):
        total_cell = sheet.cell(row=row, column=len(headers), value=counts[-1])
        total_cell.alignment = Alignment(horizontal='center')

    # Add overall total cell
    overall_total = sum(month_totals)
    overall_total_cell = sheet.cell(row=row+1, column=len(headers), value=overall_total)  
    overall_total_cell.font = Font(bold=True)
    overall_total_cell.alignment = Alignment(horizontal='center')
    
    # Apply red fill to the total column and total row
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for col in range(1, len(headers) + 1):
        sheet.cell(row=row+1, column=col).fill = red_fill
    for row in range(3, row + 1):
        sheet.cell(row=row+1, column=len(headers)).fill = red_fill

    # Autofit column width
    for col in range(1, len(headers) + 1):
        max_length = max(len(str(sheet.cell(row=row, column=col).value)) for row in range(4, total_row + 2))  # Adjusted range to include total row
        adjusted_width = max_length + 2
        sheet.column_dimensions[chr(64 + col)].width = adjusted_width

    # Return the modified sheet
    return sheet



def fetch_procedure_reports(year):
    # Get all services with the procedure category
    procedure_services = RemoteService.objects.filter(category='Procedure')

    # Query the database to get patient counts grouped by procedure category and month
    procedures_by_month = (
        RemoteProcedure.objects.filter(created_at__year=year)
        .annotate(month=ExtractMonth('created_at'))
        .values('name__name', 'month')
        .annotate(num_patients=Count('id'))
    )

    # Organize the data into a dictionary
    procedure_reports = {}
    for procedure_service in procedure_services:
        procedure_name = procedure_service.name
        procedure_reports[procedure_name] = [0] * 12  # Initialize counts for each month

    for procedure in procedures_by_month:
        procedure_name = procedure['name__name']
        month = procedure['month']
        num_patients = procedure['num_patients']

        if month is not None:
            month_index = int(month) - 1
            procedure_reports[procedure_name][month_index] = num_patients

    return procedure_reports


def render_procedure_reports(sheet, year):
    # Fetch procedure report data for the given year
    procedure_reports = fetch_procedure_reports(year)

    # Add title
    title_font = Font(size=14, bold=True, color="000000")
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = f'Site Name:'
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')
    # Add subtitle
    subtitle_font = Font(size=12, italic=True, color="808080")
    subtitle_cell = sheet.cell(row=2, column=1)
    subtitle_cell.value = 'Patient Counts by Procedure and Month'
    subtitle_cell.font = subtitle_font
    subtitle_cell.alignment = Alignment(horizontal='center')

    # Merge subtitle cells
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=14)

    # Add blank row
    blank_row_cell = sheet.cell(row=3, column=1)
    blank_row_cell.value = ''
    blank_row_cell.alignment = Alignment(horizontal='center')

    # Add headers
    headers = ['Procedure'] + [datetime.strptime(str(month), '%m').strftime('%B') for month in range(1, 13)] + ['Total']
    for col, header in enumerate(headers, start=1):
        header_cell = sheet.cell(row=4, column=col, value=header)
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center')

    # Add data
    row = 5
    for procedure, counts in procedure_reports.items():
        procedure_cell = sheet.cell(row=row, column=1, value=procedure)
        procedure_cell.alignment = Alignment(horizontal='center')
        for col, count in enumerate(counts, start=2):
            count_cell = sheet.cell(row=row, column=col, value=count)
            count_cell.alignment = Alignment(horizontal='center')
        row += 1

    # Calculate overall totals for each month
    month_totals = [sum(counts) for counts in zip(*procedure_reports.values())]

    # Add total row
    total_row_cell = sheet.cell(row=row, column=1, value='Total')
    total_row_cell.font = Font(bold=True)
    total_row_cell.alignment = Alignment(horizontal='center')
    for col, total in enumerate(month_totals, start=2):
        total_cell = sheet.cell(row=row, column=col, value=total)
        total_cell.alignment = Alignment(horizontal='center')

    # Add total column
    total_column_cell = sheet.cell(row=4, column=len(headers), value='Total')
    total_column_cell.font = Font(bold=True)
    total_column_cell.alignment = Alignment(horizontal='center')
    for row_num, (procedure, counts) in enumerate(procedure_reports.items(), start=5):
        total_cell = sheet.cell(row=row_num, column=len(headers), value=sum(counts))
        total_cell.alignment = Alignment(horizontal='center')

    # Add overall total cell
    overall_total = sum(month_totals)
    overall_total_cell = sheet.cell(row=row, column=len(headers), value=overall_total)
    overall_total_cell.font = Font(bold=True)
    overall_total_cell.alignment = Alignment(horizontal='center')

    # Apply red fill to the total column and total row
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for col in range(1, len(headers) + 1):
        sheet.cell(row=row, column=col).fill = red_fill
    for row in range(3, row + 1):
        sheet.cell(row=row, column=len(headers)).fill = red_fill

    # Autofit column width
    for col in range(1, len(headers) + 2):
        max_length = max(len(str(sheet.cell(row=row, column=col).value)) for row in range(4, row + 1))
        adjusted_width = max_length + 2
        sheet.column_dimensions[chr(64 + col)].width = adjusted_width

    return sheet


  
  
def fetch_patient_laboratory_reports(year):
    # Get all distinct laboratory names
    laboratory_services = RemoteService.objects.filter(category='Laboratory')

    # Query the database to get patient counts grouped by laboratory and month
    laboratories_by_month = (
            RemoteLaboratoryOrder.objects.filter(created_at__year=year)
            .annotate(month=ExtractMonth('created_at'))
            .values('name__name', 'month')
            .annotate(num_patients=Count('id'))
        )

    # Organize the data into a dictionary
    laboratory_reports = {}
    for laboratory_service in laboratory_services:
        laboratory_name = laboratory_service.name
        laboratory_reports[laboratory_name] = [0] * 12 
    month_totals = [0] * 12

    for laboratory in laboratories_by_month:
        laboratory_name = laboratory['name__name']
        month = laboratory['month']
        num_patients = laboratory['num_patients']

        if month is not None:
                month_index = int(month) - 1
                laboratory_reports[laboratory_name][month_index] = num_patients

    return laboratory_reports, month_totals

def render_patient_laboratory_reports(sheet, year):
    # Fetch patient laboratory-wise report data for the given year
    laboratory_reports, month_totals = fetch_patient_laboratory_reports(year)

    # Add title
    title_font = Font(size=14, bold=True, color="000000")
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = f'Site Name:'
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')
    # Add subtitle
    subtitle_font = Font(size=12, italic=True, color="808080")
    subtitle_cell = sheet.cell(row=2, column=1)
    subtitle_cell.value = 'Patient Counts by Laboratory and Month'
    subtitle_cell.font = subtitle_font
    subtitle_cell.alignment = Alignment(horizontal='center')

    # Merge subtitle cells
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=14)  # Adjust end_column

    # Add blank row
    blank_row_cell = sheet.cell(row=3, column=1)
    blank_row_cell.value = ''
    blank_row_cell.alignment = Alignment(horizontal='center')

    # Add headers
    headers = ['Laboratory'] + [datetime.strptime(str(month), '%m').strftime('%B') for month in range(1, 13)] + ['Total']  # Add 'Total' header
    for col, header in enumerate(headers, start=1):
        header_cell = sheet.cell(row=4, column=col, value=header)
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center')

    # Add data
    row = 5
    for laboratory, counts in laboratory_reports.items():
        laboratory_cell = sheet.cell(row=row, column=1, value=laboratory)
        laboratory_cell.alignment = Alignment(horizontal='center')
        for col, count in enumerate(counts, start=2):
            count_cell = sheet.cell(row=row, column=col, value=count)
            count_cell.alignment = Alignment(horizontal='center')
        row += 1
        
     # Calculate overall totals for each month
    month_totals = [sum(counts) for counts in zip(*laboratory_reports.values())]

    # Add total row
    total_row_cell = sheet.cell(row=row, column=1, value='Total')
    total_row_cell.font = Font(bold=True)
    total_row_cell.alignment = Alignment(horizontal='center')
    for col, total in enumerate(month_totals, start=2):
        total_cell = sheet.cell(row=row, column=col, value=total)
        total_cell.alignment = Alignment(horizontal='center')

    # Add total column
    total_column_cell = sheet.cell(row=4, column=len(headers), value='Total')
    total_column_cell.font = Font(bold=True)
    total_column_cell.alignment = Alignment(horizontal='center')
    for row_num, (laboratory, counts) in enumerate(laboratory_reports.items(), start=5):
        total_cell = sheet.cell(row=row_num, column=len(headers), value=sum(counts))
        total_cell.alignment = Alignment(horizontal='center')

    # Add overall total cell
    overall_total = sum(month_totals)
    overall_total_cell = sheet.cell(row=row, column=len(headers), value=overall_total)
    overall_total_cell.font = Font(bold=True)
    overall_total_cell.alignment = Alignment(horizontal='center')

    
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for col in range(1, len(headers) + 1):
        sheet.cell(row=row, column=col).fill = red_fill
    for row in range(4, row + 1):
        sheet.cell(row=row, column=len(headers)).fill = red_fill
    # Autofit column width
    for col in range(1, len(headers) + 2):  # Adjusted the range to include the overall total column
        max_length = max(len(str(sheet.cell(row=row, column=col).value)) for row in range(4, row + 1))
        adjusted_width = max_length + 2
        sheet.column_dimensions[chr(64 + col)].width = adjusted_width        
        
        
# Update render_comprehensive_report function to include pathology report
def render_comprehensive_report(year):
    wb = Workbook()
    errors = []

    # 1. Company wise report
    try:
        company_wise_sheet = wb.active
        company_wise_sheet.title = f'Consult. Per Status Report {year}'
        render_patient_company_wise_reports(company_wise_sheet, year)
    except Exception as e:
        errors.append(f"Company report error: {str(e)}")

    # 2. Pathology report
    try:
        pathology_sheet = wb.create_sheet(title=f'Consult. Per Pathology {year}')
        render_pathology_report(pathology_sheet, year)
    except Exception as e:
        errors.append(f"Pathology report error: {str(e)}")

    # 3. Procedure report
    try:
        procedure_sheet = wb.create_sheet(title=f'Nursing Procedure Report {year}')
        render_procedure_reports(procedure_sheet, year)
    except Exception as e:
        errors.append(f"Procedure report error: {str(e)}")

    # 4. Laboratory results report
    try:
        laboratory_sheet = wb.create_sheet(title=f'Laboratory Tests Report {year}')
        render_patient_laboratory_reports(laboratory_sheet, year)
    except Exception as e:
        errors.append(f"Laboratory report error: {str(e)}")

    # 5. Patient type report
    try:
        patient_type_sheet = wb.create_sheet(title=f'Patient Type Report {year}')
        render_patient_type_wise_reports(patient_type_sheet, year)
    except Exception as e:
        errors.append(f"Patient type report error: {str(e)}")

    # 6. Referral report
    try:
        referral_sheet = wb.create_sheet(title=f'Referral & MedEvac Report {year}')
        render_patient_referral_reports(referral_sheet, year)
    except Exception as e:
        errors.append(f"Referral report error: {str(e)}")

    # If any errors occurred, show them in an Excel sheet and return that
    if errors:
        error_sheet = wb.create_sheet(title='Errors')
        for i, err in enumerate(errors, start=1):
            error_sheet.cell(row=i, column=1, value=err)

    # Return the Excel file as a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="comprehensive_report_{year}.xlsx"'
    wb.save(response)
    return response


def fetch_patient_company_wise_reports(year):
    # Get all distinct company names
    all_companies = RemoteCompany.objects.values_list('name', flat=True)

    # Query the database to get patient counts grouped by company and month
    patients_by_company = (
        RemotePatient.objects.filter(created_at__year=year)
        .values('company__name')
        .annotate(month=ExtractMonth('created_at'))
        .annotate(num_patients=Count('id'))
    )

    # Organize the data into a dictionary
    company_reports = {company: [0] * 13 for company in all_companies}  # Add one more column for the total
    month_totals = [0] * 12

    for patient in patients_by_company:
        company_name = patient['company__name']
        month = patient['month']
        num_patients = patient['num_patients']

        if month is not None:
            month_index = month - 1  # ExtractMonth returns month as an integer
            company_reports[company_name][month_index] = num_patients
            company_reports[company_name][-1] += num_patients  # Add to the total column
            month_totals[month_index] += num_patients  # Add to the total row

    return company_reports, month_totals

def render_patient_company_wise_reports(sheet, year):
    # Fetch patient company-wise report data for the given year
    company_reports, month_totals = fetch_patient_company_wise_reports(year)

    # Add title
    title_font = Font(size=14, bold=True, color="000000")
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = f'Site Name:'
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')
    # Add subtitle
    subtitle_font = Font(size=12, italic=True, color="808080")
    subtitle_cell = sheet.cell(row=2, column=1)
    subtitle_cell.value = 'Patient Counts by Company and Month'
    subtitle_cell.font = subtitle_font
    subtitle_cell.alignment = Alignment(horizontal='center')

    # Merge subtitle cells
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=14)  # Adjust end_column

    # Add blank row
    blank_row_cell = sheet.cell(row=3, column=1)
    blank_row_cell.value = ''
    blank_row_cell.alignment = Alignment(horizontal='center')

    # Add headers
    headers = ['Site POB:'] + [datetime.strptime(str(month), '%m').strftime('%B') for month in range(1, 13)] + ['Total']  # Add 'Total' header
    for col, header in enumerate(headers, start=1):
        header_cell = sheet.cell(row=4, column=col, value=header)
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center')

    # Add data
    row = 5
    for company, counts in company_reports.items():
        company_cell = sheet.cell(row=row, column=1, value=company)
        company_cell.alignment = Alignment(horizontal='center')
        for col, count in enumerate(counts, start=2):
            count_cell = sheet.cell(row=row, column=col, value=count)
            count_cell.alignment = Alignment(horizontal='center')
        row += 1
        
    overall_total = sum(month_totals)
    # Add total row
    total_row_cell = sheet.cell(row=row, column=1, value='Total')
    total_row_cell.font = Font(bold=True)
    total_row_cell.alignment = Alignment(horizontal='center')
    for col, total in enumerate(month_totals, start=2):
        total_cell = sheet.cell(row=row, column=col, value=total)
        total_cell.alignment = Alignment(horizontal='center')

    # Add total column
    total_column_cell = sheet.cell(row=4, column=len(headers), value='Total')
    total_column_cell.font = Font(bold=True)
    total_column_cell.alignment = Alignment(horizontal='center')
    for row, (company, counts) in enumerate(company_reports.items(), start=5):
        total_cell = sheet.cell(row=row, column=len(headers), value=counts[-1])
        total_cell.alignment = Alignment(horizontal='center')

    
    # Add overall total cell
    overall_total_cell = sheet.cell(row=row+1, column=len(headers), value=overall_total)  # Adjusted to the last column
    overall_total_cell.font = Font(bold=True)
    overall_total_cell.alignment = Alignment(horizontal='center')
    
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for col in range(1, len(headers) + 1):
        sheet.cell(row=row+1, column=col).fill = red_fill
    for row in range(3, row + 1):
        sheet.cell(row=row+1, column=len(headers)).fill = red_fill
    # Autofit column width
    for col in range(1, len(headers) + 2):  # Adjusted the range to include the overall total column
        max_length = max(len(str(sheet.cell(row=row, column=col).value)) for row in range(4, row + 1))
        adjusted_width = max_length + 2
        sheet.column_dimensions[chr(64 + col)].width = adjusted_width

     
def fetch_patient_type_wise_reports(year):
    # Define the list of all patient types
    all_patient_types = ['National Staff', 'International Staff', 'National Visitor', 'International Visitor', 'Unknown Status', 'Others']

    # Query the database to get patient counts grouped by patient type and month
    patients_by_type = (
        RemotePatient.objects.filter(created_at__year=year)
        .values('patient_type')
        .annotate(month=ExtractMonth('created_at'))
        .annotate(num_patients=Count('id'))
    )

    # Organize the data into a dictionary
    patient_type_reports = {}   
    for patient_type in all_patient_types:
        # Initialize counts for each month
        patient_type_reports[patient_type] = [0] * 12
    month_totals = [0] * 12

    for patient in patients_by_type:
        patient_type = patient['patient_type']
        month = patient['month']
        num_patients = patient['num_patients']

        if month is not None:
            month_index = month - 1  # ExtractMonth returns month as an integer
            if patient_type not in patient_type_reports:
                patient_type_reports[patient_type] = [0] * 13  # Add one more column for the total
            patient_type_reports[patient_type][month_index] = num_patients
            patient_type_reports[patient_type][-1] += num_patients  # Add to the total column
            month_totals[month_index] += num_patients  # Add to the total row

    return patient_type_reports, month_totals

def render_patient_type_wise_reports(sheet, year):
    # Fetch patient type-wise report data for the given year
    patient_type_reports, month_totals = fetch_patient_type_wise_reports(year)

    # Add title
    title_font = Font(size=14, bold=True, color="000000")
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = f'Site Name:'
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')
    # Add subtitle
    subtitle_font = Font(size=12, italic=True, color="808080")
    subtitle_cell = sheet.cell(row=2, column=1)
    subtitle_cell.value = 'Patient Counts by Type and Month'
    subtitle_cell.font = subtitle_font
    subtitle_cell.alignment = Alignment(horizontal='center')

    # Merge subtitle cells
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=14)  # Adjust end_column

    # Add blank row
    blank_row_cell = sheet.cell(row=3, column=1)
    blank_row_cell.value = ''
    blank_row_cell.alignment = Alignment(horizontal='center')

    # Add headers
    headers = ['Patient Type'] + [datetime.strptime(str(month), '%m').strftime('%B') for month in range(1, 13)] + ['Total']  # Add 'Total' header
    for col, header in enumerate(headers, start=1):
        header_cell = sheet.cell(row=4, column=col, value=header)
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center')

    # Add data
    row = 5
    for patient_type, counts in patient_type_reports.items():
        patient_type_cell = sheet.cell(row=row, column=1, value=patient_type)
        patient_type_cell.alignment = Alignment(horizontal='center')
        for col, count in enumerate(counts, start=2):
            count_cell = sheet.cell(row=row, column=col, value=count)
            count_cell.alignment = Alignment(horizontal='center')
        row += 1
        
    overall_total = sum(month_totals)
    # Add total row
    total_row_cell = sheet.cell(row=row, column=1, value='Total')
    total_row_cell.font = Font(bold=True)
    total_row_cell.alignment = Alignment(horizontal='center')
    for col, total in enumerate(month_totals, start=2):
        total_cell = sheet.cell(row=row, column=col, value=total)
        total_cell.alignment = Alignment(horizontal='center')

    # Add total column
    total_column_cell = sheet.cell(row=4, column=len(headers), value='Total')
    total_column_cell.font = Font(bold=True)
    total_column_cell.alignment = Alignment(horizontal='center')
    for row, (patient_type, counts) in enumerate(patient_type_reports.items(), start=5):
        total_cell = sheet.cell(row=row, column=len(headers), value=counts[-1])
        total_cell.alignment = Alignment(horizontal='center')

    # Add overall total cell
    overall_total_cell = sheet.cell(row=row+1, column=len(headers), value=overall_total)  # Adjusted to the last column
    overall_total_cell.font = Font(bold=True)
    overall_total_cell.alignment = Alignment(horizontal='center')
    
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for col in range(1, len(headers) + 1):
        sheet.cell(row=row+1, column=col).fill = red_fill
    for row in range(3, row + 1):
        sheet.cell(row=row+1, column=len(headers)).fill = red_fill
    # Autofit column width
    for col in range(1, len(headers) + 2):  # Adjusted the range to include the overall total column
        max_length = max(len(str(sheet.cell(row=row, column=col).value)) for row in range(4, row + 1))
        adjusted_width = max_length + 2
        sheet.column_dimensions[chr(64 + col)].width = adjusted_width
        
        
def fetch_patient_referral_reports(year):
    # Fetch data for patient referral report
    referrals = RemoteReferral.objects.filter(created_at__year=year)
    return referrals

def render_patient_referral_reports(sheet, year):
    # Fetch patient referral report data for the given year
    referrals = fetch_patient_referral_reports(year)

    # Add title
    title_font = Font(size=14, bold=True, color="000000")
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = f'Site Name:'
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')
    # Add headers
    headers = ['Date', 'Patient mrn', 'First Name', 'Family Name', 'Age', 'Sex', 'Nationality',
               'Company', 'Patient Category', 'Med Evac/Refererred', 'Referral Reason', 'Transport Mode',
               'Patient Destination', 'Diagnosis']

    for col, header in enumerate(headers, start=1):
        header_cell = sheet.cell(row=2, column=col)  # Note: Starting from row 2
        header_cell.value = header
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center')

    # Add data
    for idx, referral in enumerate(referrals, start=3):  # Note: Starting from row 3
        age = referral.patient.age if referral.patient.dob else ''
        if referral.patient.dob:
            now = datetime.now()
            age = now.year - referral.patient.dob.year - ((now.month, now.day) < (referral.patient.dob.month, referral.patient.dob.day))

        diagnosis_content = ''
        for diagnosis_record in referral.patient.remotepatientdiagnosisrecord_set.all():
            for diagnosis in diagnosis_record.final_diagnosis.all():
                diagnosis_content += f'{diagnosis}, '

        # Parse notes
        notes_content = referral.notes.replace('<ol>', '').replace('<li>', '- ').replace('</li>', '\n').replace('</ol>', '')
        # Append referral data to Excel sheet
        sheet.cell(row=idx, column=1).value = referral.created_at.strftime('%d/%m/%Y')
        sheet.cell(row=idx, column=2).value = referral.patient.mrn
        sheet.cell(row=idx, column=3).value = referral.patient.first_name
        sheet.cell(row=idx, column=4).value = referral.patient.last_name
        sheet.cell(row=idx, column=5).value = f'{age} years'
        sheet.cell(row=idx, column=6).value = referral.patient.gender
        sheet.cell(row=idx, column=7).value = referral.patient.nationality.name
        sheet.cell(row=idx, column=8).value = referral.patient.company.name
        sheet.cell(row=idx, column=9).value = referral.patient.patient_type
        sheet.cell(row=idx, column=10).value = referral.nature_of_referral
        sheet.cell(row=idx, column=11).value = notes_content
        sheet.cell(row=idx, column=12).value = referral.transport_model        
        sheet.cell(row=idx, column=13).value = referral.destination_location
        sheet.cell(row=idx, column=14).value = diagnosis_content.rstrip(', ')

    # Autofit column width
    for col in range(1, 15):
        max_length = max(len(str(sheet.cell(row=row, column=col).value)) for row in range(2, len(referrals) + 3))
        adjusted_width = max_length + 2
        sheet.column_dimensions[chr(64 + col)].width = adjusted_width

    # Save the workbook with the report
    return sheet

        
        
def generate_year_month_report(request):
    try:
        if request.method == 'POST':
            form = YearMonthSelectionForm(request.POST)

            if form.is_valid():
                year = int(form.cleaned_data['year'])
                month = form.cleaned_data['month']  # Could be 0 or None

                # Yearly Report
                if month == 0 or month is None:
                    try:
                        response = render_comprehensive_report(year)
                        messages.success(request, f"Yearly report for {year} generated successfully.")
                        return response
                    except Exception as e:
                        messages.error(request, f"Failed to generate yearly report for {year}. Reason: {str(e)}")
                
                # Monthly Report
                else:
                    try:
                        response = render_daily_comprehensive_report(year, month)
                        messages.success(request, f"Monthly report for {year}-{month:02} generated successfully.")
                        return response
                    except Exception as e:
                        messages.error(request, f"Failed to generate monthly report for {year}-{month:02}. Reason: {str(e)}")

            else:
                messages.error(request, "Invalid input: Please select a valid year and month.")
        else:
            form = YearMonthSelectionForm()

        return render(request, 'divine_admin_template/generate_year_month_report.html', {'form': form})

    except Exception as e:
        print(traceback.format_exc())  # Or use logging
        messages.error(request, f"An unexpected error occurred: {str(e)}")
        form = YearMonthSelectionForm()
        return render(request, 'divine_admin_template/generate_year_month_report.html', {'form': form})



def render_daily_comprehensive_report(year, month):
    # Create a new Excel workbook
    wb = Workbook()

    # Safe month name
    safe_month_name = calendar.month_name[month].replace("/", "-").replace("[", "").replace("]", "")

    # Add company-wise report
    company_wise_sheet = wb.active
    company_wise_sheet.title = f'Consult Status {safe_month_name} {year}'[:31]
    render_daily_patient_company_wise_reports(company_wise_sheet, year, month)

    # # Add pathology report
    pathology_sheet = wb.create_sheet(title=f'Consult Pathology {safe_month_name} {year}'[:31])
    render_daily_pathology_report(pathology_sheet, year, month)

    # # Add procedure report
    procedure_sheet = wb.create_sheet(title=f'Nursing Procedure {safe_month_name} {year}'[:31])
    render_daily_procedure_reports(procedure_sheet, year, month)

    # # Add laboratory result report
    laboratory_sheet = wb.create_sheet(title=f'Lab Tests {safe_month_name} {year}'[:31])
    render_daily_patient_laboratory_reports(laboratory_sheet, year, month)

    # # Add patient type report
    patient_type_sheet = wb.create_sheet(title=f'Patient Type {safe_month_name} {year}'[:31])
    render_daily_patient_type_wise_reports(patient_type_sheet, year, month)

    # # Add patient referral report
    referral_sheet = wb.create_sheet(title=f'Referral MedEvac {safe_month_name} {year}'[:31])
    render_daily_patient_referral_reports(referral_sheet, year, month)

    # Prepare response to return the Excel workbook
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="daily_report_{safe_month_name}_{year}.xlsx"'
    wb.save(response)

    return response



def render_daily_patient_type_wise_reports(patient_type_sheet, year, month):
    # Add title
    title_font = Font(size=14, bold=True, color="000000")
    title_cell = patient_type_sheet.cell(row=1, column=1)
    title_cell.value = f'Site Name: '
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')

    # Add subtitle
    subtitle_font = Font(size=12, italic=True, color="808080")
    subtitle_cell = patient_type_sheet.cell(row=2, column=1)
    subtitle_cell.value = f'Patient type and Date for {calendar.month_name[month]} {year}'
    subtitle_cell.font = subtitle_font
    subtitle_cell.alignment = Alignment(horizontal='center')

    # Fetch daily data
    patient_type_reports, day_totals = fetch_daily_patient_type_wise_reports(year, month)

    # Generate dates for the given month (1st to 31st)
    num_days_in_month = monthrange(year, month)[1]
    dates = [f'{year}-{month:02d}-{day:02d}' for day in range(1, num_days_in_month + 1)]

    # Adjust the row number after title and subtitle
    row_num = 4  # Start data rows after title and subtitle

    # Add headers
    headers = ['Date'] + list(patient_type_reports.keys()) + ['Total']
    for col, header in enumerate(headers, start=1):
        cell = patient_type_sheet.cell(row=row_num, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Add data rows
    row_num += 1  # Move to the next row after headers
    column_totals = [0] * (len(patient_type_reports) + 1)  # Initialize column totals
    for idx, date in enumerate(dates):
        patient_type_sheet.cell(row=row_num, column=1, value=date)  # Add date to the first column
        col_num = 2
        for patient_type, counts in patient_type_reports.items():
            patient_type_sheet.cell(row=row_num, column=col_num, value=counts[idx])  # Add patient count for each type
            column_totals[col_num - 2] += counts[idx]  # Accumulate column totals
            col_num += 1
        patient_type_sheet.cell(row=row_num, column=col_num, value=day_totals[idx])  # Add the total for this day
        row_num += 1

    # Apply red color to the last column (Total column) and last row (Total row)
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Add the column totals
    for col, total in enumerate(column_totals, start=2):
        cell = patient_type_sheet.cell(row=row_num, column=col, value=total)
        cell.font = Font(bold=True)
        cell.fill = red_fill  # Apply red fill to the total column

    # Apply red fill to the last cell (Total of all columns)
    overall_total = sum(day_totals)
    total_col_num = len(headers)
    total_cell = patient_type_sheet.cell(row=row_num, column=total_col_num, value=overall_total)
    total_cell.font = Font(bold=True)
    total_cell.fill = red_fill  # Apply red fill to the overall total column

    # Autofit columns
    for col in patient_type_sheet.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value) + 2
        col_letter = col[0].column_letter
        patient_type_sheet.column_dimensions[col_letter].width = max_length

    return patient_type_sheet
        
        
def fetch_daily_patient_type_wise_reports(year, month):
    # Define the list of all patient types
    all_patient_types = ['National Staff', 'International Staff', 'National Visitor', 'International Visitor', 'Unknown Status', 'Others']

    # Query the database to get patient counts grouped by patient type and day
    patients_by_type = (
        RemotePatient.objects.filter(created_at__year=year, created_at__month=month)
        .annotate(day=TruncDay('created_at'))
        .values('day', 'patient_type')
        .annotate(count=Count('id'))  # The count is stored under the key 'count'
    )

    # Organize the data into a dictionary
    patient_type_reports = {}
    for patient_type in all_patient_types:
        # Initialize counts for each day (1-31)
        patient_type_reports[patient_type] = [0] * 31
    day_totals = [0] * 31

    for patient in patients_by_type:
        patient_type = patient['patient_type']
        day = patient['day']
        num_patients = patient['count']  # Use 'count' instead of 'num_patients'

        if day is not None:
            day_index = day.day - 1  # Extract the day as an integer
            if patient_type not in patient_type_reports:
                patient_type_reports[patient_type] = [0] * 32  # Add one more column for the total
            patient_type_reports[patient_type][day_index] = num_patients
            day_totals[day_index] += num_patients  # Add to the total row

    return patient_type_reports, day_totals

     

def fetch_daily_patient_company_wise_reports(year, month):
    # Get all distinct company names
    all_companies = RemoteCompany.objects.values_list('name', flat=True)

    # Query the database to get patient counts grouped by company and day
    patients_by_company = (
        RemotePatient.objects.filter(created_at__year=year, created_at__month=month)
        .values('company__name')
        .annotate(day=TruncDay('created_at'))
        .annotate(num_patients=Count('id'))
    )

    # Organize the data into a dictionary
    company_reports = {company: [0] * 31 for company in all_companies}  # Initialize 31 days
    day_totals = [0] * 31

    for patient in patients_by_company:
        company_name = patient['company__name']
        day = patient['day']
        num_patients = patient['num_patients']

        if day is not None:
            day_index = day.day - 1  # ExtractDay returns day as an integer
            company_reports[company_name][day_index] = num_patients
            day_totals[day_index] += num_patients  # Add to the total row

    return company_reports, day_totals


def render_daily_patient_company_wise_reports(sheet, year, month):
    # Add title
    title_font = Font(size=14, bold=True, color="000000")
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = f'Site Name: '
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')

    # Add subtitle
    subtitle_font = Font(size=12, italic=True, color="808080")
    subtitle_cell = sheet.cell(row=2, column=1)
    subtitle_cell.value = f'Patient Counts by company Record and Date for {calendar.month_name[month]} {year}'
    subtitle_cell.font = subtitle_font
    subtitle_cell.alignment = Alignment(horizontal='center')

    # Fetch patient company-wise report data for the given year and month
    company_reports, day_totals = fetch_daily_patient_company_wise_reports(year, month)

    # Add headers
    headers = ['Date'] + list(company_reports.keys()) + ['Total']
    row_num = 3  # Start from row 3 after title and subtitle
    for col, header in enumerate(headers, start=1):
        header_cell = sheet.cell(row=row_num, column=col, value=header)
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center')

    # Add data rows (dates)
    num_days = calendar.monthrange(year, month)[1]  # Get the number of days in the month
    for day in range(1, num_days + 1):
        date_value = datetime(year, month, day).strftime('%Y-%m-%d')
        row_num = day + 3  # Adjust row number to start from row 4 for data rows
        sheet.cell(row=row_num, column=1, value=date_value)  # Add date to the first column
        col_num = 2
        for company, counts in company_reports.items():
            sheet.cell(row=row_num, column=col_num, value=counts[day - 1])
            col_num += 1
        sheet.cell(row=row_num, column=col_num, value=day_totals[day - 1])  # Total column

    # Add totals row at the bottom
    total_label = "Total"
    row_num = num_days + 4  # Total row at the end
    sheet.cell(row=row_num, column=1, value=total_label)
    sheet.cell(row=row_num, column=1).font = Font(bold=True)
    for col, company in enumerate(company_reports.keys(), start=2):
        total = sum(company_reports[company])
        sheet.cell(row=row_num, column=col, value=total)
        sheet.cell(row=row_num, column=col).font = Font(bold=True)

    overall_total = sum(day_totals)
    total_col_num = len(headers)
    sheet.cell(row=row_num, column=total_col_num, value=overall_total)
    sheet.cell(row=row_num, column=total_col_num).font = Font(bold=True)

    # Autofit column widths
    for col in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        col_letter = col[0].column_letter
        sheet.column_dimensions[col_letter].width = max_length + 2

    return sheet
        

def fetch_daily_pathology_reports(year, month):
    # Get all distinct Pathology Record names for the given year and month
    all_pathology_records = PathodologyRecord.objects.values_list('name', flat=True)

    # Query the database to get patient counts grouped by Pathology Record and date for the given year and month
    patients_by_pathology_record = (
        PathodologyRecord.objects.filter(remoteconsultationnotes__created_at__year=year, remoteconsultationnotes__created_at__month=month)
        .annotate(day=TruncDay('remoteconsultationnotes__created_at'))
        .values('name', 'day')
        .annotate(num_patients=Count('remoteconsultationnotes__id'))
    )

    # Organize the data into a dictionary
    days_in_month = calendar.monthrange(year, month)[1]
    pathology_record_reports = {record: [0] * (days_in_month + 2) for record in all_pathology_records}  # Add one column for total
    day_totals = [0] * days_in_month

    for patient in patients_by_pathology_record:
        pathology_record_name = patient['name']
        day = patient['day']
        num_patients = patient['num_patients']

        if day is not None:
            day_index = day.day - 1  # ExtractDay returns day as an integer
            pathology_record_reports[pathology_record_name][day_index] = num_patients
            pathology_record_reports[pathology_record_name][-1] += num_patients  # Add to the total column
            day_totals[day_index] += num_patients  # Add to the total row

    return pathology_record_reports, day_totals


def render_daily_pathology_report(sheet, year, month):
    # Fetch daily pathology report data for the given year and month
    pathology_reports, day_totals = fetch_daily_pathology_reports(year, month)

    # Add title
    title_font = Font(size=14, bold=True, color="000000")
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = f'Site Name:'
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')

    # Add subtitle
    subtitle_font = Font(size=12, italic=True, color="808080")
    subtitle_cell = sheet.cell(row=2, column=1)
    subtitle_cell.value = f'Patient Counts by Pathology Record and Date for {calendar.month_name[month]} {year}'
    subtitle_cell.font = subtitle_font
    subtitle_cell.alignment = Alignment(horizontal='center')

    # Merge subtitle cells
    days_in_month = calendar.monthrange(year, month)[1]
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=days_in_month + 2)

    # Add blank row
    blank_row_cell = sheet.cell(row=3, column=1)
    blank_row_cell.value = ''
    blank_row_cell.alignment = Alignment(horizontal='center')

    # Add headers with dates instead of day numbers
    headers = ['CONSULTATIONS / PATHOLOGY'] + [str(datetime(year, month, day).date()) for day in range(1, days_in_month + 1)] + ['Total']
    for col, header in enumerate(headers, start=1):
        header_cell = sheet.cell(row=4, column=col, value=header)
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center')

    # Add data and calculate daily totals
    row = 5
    total_counts_by_day = [0] * days_in_month  # List to store the total for each day across all pathology records

    for pathology_name, counts_by_day in pathology_reports.items():
        record_cell = sheet.cell(row=row, column=1, value=pathology_name)
        record_cell.alignment = Alignment(horizontal='left')

        row_total = sum(counts_by_day)  # Total for the current pathology record
        for col, count in enumerate(counts_by_day, start=2):
            count_cell = sheet.cell(row=row, column=col, value=count)
            count_cell.alignment = Alignment(horizontal='center')
            if col - 2 < len(total_counts_by_day):  # Ensure the index is valid
                total_counts_by_day[col - 2] += count  # Add the count to the total for the corresponding day

        # Add the total for the current pathology record to the last column
        total_cell = sheet.cell(row=row, column=days_in_month + 2, value=row_total)
        total_cell.alignment = Alignment(horizontal='center')

        row += 1

    # Add total row for each day
    total_row = row
    total_row_cell = sheet.cell(row=row, column=1, value='Total')
    total_row_cell.font = Font(bold=True)
    total_row_cell.alignment = Alignment(horizontal='center')

    for col, total in enumerate(total_counts_by_day, start=2):
        total_cell = sheet.cell(row=row, column=col, value=total)
        total_cell.alignment = Alignment(horizontal='center')

    # Add overall total cell (total of all days)
    overall_total = sum(total_counts_by_day)
    overall_total_cell = sheet.cell(row=row, column=days_in_month + 2, value=overall_total)
    overall_total_cell.font = Font(bold=True)
    overall_total_cell.alignment = Alignment(horizontal='center')

    # Apply red fill to the total column and total row
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Apply red fill to total for each pathology record
    for row_index in range(5, row):
        total_cell = sheet.cell(row=row_index, column=days_in_month + 2)
        total_cell.fill = red_fill

    # Apply red fill to the last column of the total row
    for col in range(1, days_in_month + 3):
        total_cell = sheet.cell(row=row, column=col)
        total_cell.fill = red_fill

    # Autofit column width using openpyxl's get_column_letter for handling columns beyond Z
    for col in range(1, days_in_month + 3):
        max_length = max(len(str(sheet.cell(row=r, column=col).value)) for r in range(4, row))
        adjusted_width = max_length + 2
        col_letter = get_column_letter(col)
        sheet.column_dimensions[col_letter].width = adjusted_width

    # Return the modified sheet
    return sheet




def fetch_daily_procedure_reports(year, month):
    # Get all services with the procedure category
    procedure_services = RemoteService.objects.filter(category='Procedure')

    # Query the database to get patient counts grouped by procedure and date for the given month and year
    procedures_by_date = (
        RemoteProcedure.objects.filter(created_at__year=year, created_at__month=month)
        .annotate(day=TruncDay('created_at'))
        .values('name__name', 'created_at')
        .annotate(num_patients=Count('id'))
    )

    # Organize the data into a dictionary
    days_in_month = calendar.monthrange(year, month)[1]
    procedure_reports = {}
    for procedure_service in procedure_services:
        procedure_name = procedure_service.name
        procedure_reports[procedure_name] = [0] * days_in_month  # Initialize counts for each day

    # Populate procedure reports with patient counts for each date
    for procedure in procedures_by_date:
        procedure_name = procedure['name__name']
        date = procedure['created_at']
        num_patients = procedure['num_patients']

        if date is not None:
            # Find the day of the month for the given date
            day_index = date.day - 1
            procedure_reports[procedure_name][day_index] = num_patients

    return procedure_reports


def render_daily_procedure_reports(sheet, year, month):
    # Fetch daily procedure report data for the given year and month
    procedure_reports = fetch_daily_procedure_reports(year, month)

    # Add title
    title_font = Font(size=14, bold=True, color="000000")
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = f'Site Name:'
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')

    # Add subtitle
    subtitle_font = Font(size=12, italic=True, color="808080")
    subtitle_cell = sheet.cell(row=2, column=1)
    subtitle_cell.value = f'Patient Counts by Procedure and Date for {calendar.month_name[month]} {year}'
    subtitle_cell.font = subtitle_font
    subtitle_cell.alignment = Alignment(horizontal='center')

    # Merge subtitle cells
    days_in_month = calendar.monthrange(year, month)[1]
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=days_in_month + 2)

    # Add blank row
    blank_row_cell = sheet.cell(row=3, column=1)
    blank_row_cell.value = ''
    blank_row_cell.alignment = Alignment(horizontal='center')

    # Add headers
    headers = ['Procedure'] + [str(datetime(year, month, day).date()) for day in range(1, days_in_month + 1)] + ['Total']
    for col, header in enumerate(headers, start=1):
        header_cell = sheet.cell(row=4, column=col, value=header)
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center')

    # Add data and calculate totals
    row = 5
    total_counts = [0] * days_in_month  # List to store the total for each day across all procedures

    for procedure, counts in procedure_reports.items():
        procedure_cell = sheet.cell(row=row, column=1, value=procedure)
        procedure_cell.alignment = Alignment(horizontal='center')

        row_total = sum(counts)  # Total for the current procedure
        for col, count in enumerate(counts, start=2):
            count_cell = sheet.cell(row=row, column=col, value=count)
            count_cell.alignment = Alignment(horizontal='center')
            total_counts[col - 2] += count  # Add the count to the total for the corresponding day

        # Add the total for the current procedure to the last column
        total_cell = sheet.cell(row=row, column=days_in_month + 2, value=row_total)
        total_cell.alignment = Alignment(horizontal='center')

        row += 1

    # Add overall total row for all days
    total_row_cell = sheet.cell(row=row, column=1, value='Total')
    total_row_cell.font = Font(bold=True)
    total_row_cell.alignment = Alignment(horizontal='center')

    for col, total in enumerate(total_counts, start=2):
        total_cell = sheet.cell(row=row, column=col, value=total)
        total_cell.alignment = Alignment(horizontal='center')

    # Add overall total cell (total of all days)
    overall_total = sum(total_counts)
    overall_total_cell = sheet.cell(row=row, column=days_in_month + 2, value=overall_total)
    overall_total_cell.font = Font(bold=True)
    overall_total_cell.alignment = Alignment(horizontal='center')

    # Apply red fill to the total column and total row
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for col in range(1, days_in_month + 3):
        sheet.cell(row=row, column=col).fill = red_fill
    for row_num in range(4, row + 1):
        sheet.cell(row=row_num, column=days_in_month + 2).fill = red_fill

    # Autofit column width using openpyxl's get_column_letter for handling columns beyond Z
    for col in range(1, days_in_month + 3):
        max_length = max(len(str(sheet.cell(row=row_num, column=col).value)) for row_num in range(4, row + 1))
        adjusted_width = max_length + 2
        col_letter = get_column_letter(col)  # Use get_column_letter for correct column naming
        sheet.column_dimensions[col_letter].width = adjusted_width

    return sheet


def fetch_daily_patient_laboratory_reports(year, month):
    # Get all distinct laboratory names
    laboratory_services = RemoteService.objects.filter(category='Laboratory')

    # Query the database to get patient counts grouped by laboratory and date for the given month and year
    laboratories_by_date = (
        RemoteLaboratoryOrder.objects.filter(created_at__year=year, created_at__month=month)
        .annotate(day=TruncDay('created_at'))  # Truncates to the date part (year-month-day)
        .values('name__name', 'created_at')  # Use date instead of day
        .annotate(num_patients=Count('id'))
    )

    # Organize the data into a dictionary
    days_in_month = calendar.monthrange(year, month)[1]
    laboratory_reports = {}
    for laboratory_service in laboratory_services:
        laboratory_name = laboratory_service.name
        laboratory_reports[laboratory_name] = [0] * days_in_month  # Initialize counts for each day

    # Populate the dictionary with actual data from the database
    for laboratory in laboratories_by_date:
        laboratory_name = laboratory['name__name']
        date = laboratory['created_at']  # Get date from the TruncDate result
        num_patients = laboratory['num_patients']

        if date is not None:
            day_index = date.day - 1  # Index for the days (0-based index)
            laboratory_reports[laboratory_name][day_index] = num_patients

    return laboratory_reports


def render_daily_patient_laboratory_reports(sheet, year, month):
    # Fetch daily laboratory report data for the given year and month
    laboratory_reports = fetch_daily_patient_laboratory_reports(year, month)

    # Add title
    title_font = Font(size=14, bold=True, color="000000")
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = f'Site Name:'
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')

    # Add subtitle
    subtitle_font = Font(size=12, italic=True, color="808080")
    subtitle_cell = sheet.cell(row=2, column=1)
    subtitle_cell.value = f'Patient Counts by Laboratory and Date for {calendar.month_name[month]} {year}'
    subtitle_cell.font = subtitle_font
    subtitle_cell.alignment = Alignment(horizontal='center')

    # Merge subtitle cells
    days_in_month = calendar.monthrange(year, month)[1]
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=days_in_month + 2)

    # Add blank row
    blank_row_cell = sheet.cell(row=3, column=1)
    blank_row_cell.value = ''
    blank_row_cell.alignment = Alignment(horizontal='center')

    # Add headers with actual dates
    headers = ['Laboratory'] + [
        (datetime(year, month, day).strftime('%Y-%m-%d')) for day in range(1, days_in_month + 1)
    ] + ['Total']
    
    for col, header in enumerate(headers, start=1):
        header_cell = sheet.cell(row=4, column=col, value=header)
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center')

    # Add data
    row = 5
    laboratory_totals = []  # To store totals for each laboratory
    for laboratory, counts in laboratory_reports.items():
        laboratory_cell = sheet.cell(row=row, column=1, value=laboratory)
        laboratory_cell.alignment = Alignment(horizontal='center')
        
        total_for_laboratory = 0  # To calculate the total per laboratory
        for col, count in enumerate(counts, start=2):
            count_cell = sheet.cell(row=row, column=col, value=count)
            count_cell.alignment = Alignment(horizontal='center')
            total_for_laboratory += count  # Sum up the counts for the laboratory
        
        # Add the total for the laboratory to the last column
        total_cell = sheet.cell(row=row, column=days_in_month + 2, value=total_for_laboratory)
        total_cell.alignment = Alignment(horizontal='center')
        
        laboratory_totals.append(total_for_laboratory)  # Store the laboratory total
        row += 1

    # Calculate overall totals for each date
    date_totals = [sum(counts) for counts in zip(*laboratory_reports.values())]

    # Add total row
    total_row_cell = sheet.cell(row=row, column=1, value='Total')
    total_row_cell.font = Font(bold=True)
    total_row_cell.alignment = Alignment(horizontal='center')
    
    for col, total in enumerate(date_totals, start=2):
        total_cell = sheet.cell(row=row, column=col, value=total)
        total_cell.alignment = Alignment(horizontal='center')
    
    # Add overall total cell (total of all dates)
    overall_total = sum(laboratory_totals)
    overall_total_cell = sheet.cell(row=row, column=days_in_month + 2, value=overall_total)
    overall_total_cell.font = Font(bold=True)
    overall_total_cell.alignment = Alignment(horizontal='center')

    # Apply red fill to the total column and total row
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for col in range(1, days_in_month + 3):
        sheet.cell(row=row, column=col).fill = red_fill
    for row_num in range(4, row + 1):
        sheet.cell(row=row_num, column=days_in_month + 2).fill = red_fill

    # Autofit column width using openpyxl's get_column_letter for handling columns beyond Z
    for col in range(1, days_in_month + 3):
        max_length = max(len(str(sheet.cell(row=row_num, column=col).value)) for row_num in range(4, row + 1))
        adjusted_width = max_length + 2
        col_letter = get_column_letter(col)  # Use get_column_letter for correct column naming
        sheet.column_dimensions[col_letter].width = adjusted_width

    return sheet



def fetch_daily_patient_referral_reports(year, month):
    # Fetch data for patient referral report filtered by the specific month and year
    referrals = RemoteReferral.objects.filter(created_at__year=year, created_at__month=month)
    return referrals

def render_daily_patient_referral_reports(sheet, year, month):
    # Fetch patient referral report data for the given year and month
    referrals = fetch_daily_patient_referral_reports(year, month)

    # Add title
    title_font = Font(size=14, bold=True, color="000000")
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = f'Site Name:'
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')

    # Add subtitle with the specific month and year
    subtitle_font = Font(size=12, italic=True, color="808080")
    subtitle_cell = sheet.cell(row=2, column=1)
    subtitle_cell.value = f'Patient Referral Report for {calendar.month_name[month]} {year}'
    subtitle_cell.font = subtitle_font
    subtitle_cell.alignment = Alignment(horizontal='center')

    # Merge subtitle cells to span the width of the sheet
    days_in_month = calendar.monthrange(year, month)[1]
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=days_in_month + 2)

    # Add blank row
    blank_row_cell = sheet.cell(row=3, column=1)
    blank_row_cell.value = ''
    blank_row_cell.alignment = Alignment(horizontal='center')

    # Add headers
    headers = ['Date', 'Patient MRN', 'First Name', 'Family Name', 'Age', 'Sex', 'Nationality',
               'Company', 'Patient Category', 'Med Evac/Referrals', 'Referral Reason', 'Transport Mode',
               'Patient Destination', 'Diagnosis']
    for col, header in enumerate(headers, start=1):
        header_cell = sheet.cell(row=4, column=col)
        header_cell.value = header
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center')

    # Add data for each referral
    row = 5
    for referral in referrals:
        age = referral.patient.age if referral.patient.dob else ''
        if referral.patient.dob:
            now = datetime.now()
            age = now.year - referral.patient.dob.year - ((now.month, now.day) < (referral.patient.dob.month, referral.patient.dob.day))

        # Prepare diagnosis and notes content
        diagnosis_content = ''
        for diagnosis_record in referral.patient.remotepatientdiagnosisrecord_set.all():
            for diagnosis in diagnosis_record.final_diagnosis.all():
                diagnosis_content += f'{diagnosis}, '

        notes_content = referral.notes.replace('<ol>', '').replace('<li>', '- ').replace('</li>', '\n').replace('</ol>', '')

        # Add referral data to Excel sheet
        sheet.cell(row=row, column=1).value = referral.created_at.strftime('%d/%m/%Y')  # Date of referral
        sheet.cell(row=row, column=2).value = referral.patient.mrn  # MRN
        sheet.cell(row=row, column=3).value = referral.patient.first_name  # First name
        sheet.cell(row=row, column=4).value = referral.patient.last_name  # Family name
        sheet.cell(row=row, column=5).value = f'{age} years'  # Age
        sheet.cell(row=row, column=6).value = referral.patient.gender  # Sex
        sheet.cell(row=row, column=7).value = referral.patient.nationality.name  # Nationality
        sheet.cell(row=row, column=8).value = referral.patient.company.name  # Company
        sheet.cell(row=row, column=9).value = referral.patient.patient_type  # Patient Category
        sheet.cell(row=row, column=10).value = referral.nature_of_referral  # Referral Reason
        sheet.cell(row=row, column=11).value = notes_content  # Notes
        sheet.cell(row=row, column=12).value = referral.transport_model  # Transport Mode
        sheet.cell(row=row, column=13).value = referral.destination_location  # Patient Destination
        sheet.cell(row=row, column=14).value = diagnosis_content.rstrip(', ')  # Diagnosis

        row += 1

    # Autofit column width
    for col in range(1, 15):
        max_length = max(len(str(sheet.cell(row=row_num, column=col).value)) for row_num in range(4, row + 1))
        adjusted_width = max_length + 2
        sheet.column_dimensions[chr(64 + col)].width = adjusted_width

    # Return the sheet with the generated report
    return sheet