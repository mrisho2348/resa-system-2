import openpyxl
from django.shortcuts import render
from django.http import HttpResponseRedirect
from django.urls import reverse
from django.contrib import messages
from clinic.models import Country, Diagnosis, DiseaseRecode, HealthRecord,  PathodologyRecord, RemoteCompany, RemoteEquipment, RemoteMedicine, RemoteReagent, RemoteService
from .forms import CountryImportForm, DiagnosisImportForm, DiseaseRecodeImportForm, HealthRecordImportForm, ImportInsuranceCompanyForm, PathodologyRecordImportForm, RemoteCompanyImportForm, RemoteEquipmentForm, RemoteMedicineImportForm, RemoteReagentForm,  RemoteServiceImportForm
from django.db.utils import IntegrityError





def import_disease_recode_data(request):
    if request.method == 'POST':
        form = DiseaseRecodeImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            try:
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active

                # Read headers
                headers = [cell.value.strip() for cell in sheet[1]]  # Trim headers
                required_headers = ['disease_name', 'code']

                # Validate headers
                if headers[:len(required_headers)] != required_headers:
                    messages.error(request, 'Invalid file format')
                    return render(request, 'divineImport/import_disease_recode.html', {'form': form})

                # Read data from rows
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    try:
                        # Ensure each header is mapped to a value, defaulting to an empty string if necessary
                        data = dict(zip(headers, (cell.strip() if isinstance(cell, str) else cell for cell in row)))  # Trim data
                        DiseaseRecode.objects.create(
                            disease_name=data['disease_name'],
                            code=data['code']
                        )
                    except IntegrityError:
                        # Skip duplicate entries and continue
                        continue
                    except Exception as e:
                        messages.error(request, f"Failed to import row data: {str(e)}")
                        continue

                return HttpResponseRedirect(reverse('divine_manage_disease'))

            except Exception as e:
                messages.error(request, f"Failed to import data: {str(e)}")
                return render(request, 'divineImport/import_disease_recode.html', {'form': form})
    else:
        form = DiseaseRecodeImportForm()

    return render(request, 'divineImport/import_disease_recode.html', {'form': form})


def import_remote_medicine_data(request):
    if request.method == 'POST':
        form = RemoteMedicineImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            try:
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active

                # Read headers
                headers = [cell.value.strip() for cell in sheet[1]]  # Trim headers
                required_headers = [
                    'drug_name', 'drug_type', 'formulation_unit', 'manufacturer',
                    'quantity', 'dividable', 'batch_number', 'expiration_date', 
                    'unit_cost', 'buying_price'
                ]

                # Validate headers
                if headers[:len(required_headers)] != required_headers:
                    messages.error(request, 'Invalid file format')
                    return render(request, 'divineImport/import_remote_medicine.html', {'form': form})

                # Read data from rows
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    try:
                        # Ensure each header is mapped to a value, defaulting to an empty string if necessary
                        data = dict(zip(headers, (cell.strip() if isinstance(cell, str) else cell for cell in row)))  # Trim data
                        
                        RemoteMedicine.objects.create(
                            drug_name=data['drug_name'],
                            drug_type=data['drug_type'],
                            formulation_unit=data['formulation_unit'],
                            manufacturer=data['manufacturer'],
                            quantity=data['quantity'],
                            dividable=data['dividable'],
                            batch_number=data['batch_number'],
                            expiration_date=data['expiration_date'],
                            unit_cost=data['unit_cost'],
                            buying_price=data['buying_price'],
                            total_buying_price=float(data['buying_price']) * data['quantity'] if data['buying_price'] and data['quantity'] else None,
                            remain_quantity=data['quantity'],
                        )
                    except IntegrityError:
                        # Skip duplicate entries and continue
                        continue
                    except Exception as e:
                        messages.error(request, f"Failed to import row data: {str(e)}")
                        continue

                return HttpResponseRedirect(reverse('divine_remotemedicine_list'))

            except Exception as e:
                messages.error(request, f"Failed to import data: {str(e)}")
                return render(request, 'divineImport/import_remote_medicine.html', {'form': form})
    else:
        form = RemoteMedicineImportForm()
    return render(request, 'divineImport/import_remote_medicine.html', {'form': form})



def import_health_record_data(request):
    if request.method == 'POST':
        form = HealthRecordImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            try:
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active

                # Read headers
                headers = [cell.value.strip() for cell in sheet[1]]  # Trim headers
                required_headers = ['name']

                # Validate headers
                if headers[:len(required_headers)] != required_headers:
                    messages.error(request, 'Invalid file format')
                    return render(request, 'divineImport/import_health_record.html', {'form': form})

                # Read data from rows
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    try:
                        # Ensure each header is mapped to a value, defaulting to an empty string if necessary
                        data = dict(zip(headers, (cell.strip() if isinstance(cell, str) else cell for cell in row)))  # Trim data
                        name = data.get('name')

                        if name:  # Only proceed if name is present
                            # Use get_or_create to avoid duplicates
                            HealthRecord.objects.get_or_create(
                                name=name.strip()  # Trim name
                            )
                    except IntegrityError:
                        # Skip duplicate entries and continue
                        continue
                    except Exception as e:
                        messages.error(request, f"Failed to import row data: {str(e)}")
                        continue

                return HttpResponseRedirect(reverse('divine_health_record_list'))

            except Exception as e:
                messages.error(request, f"Failed to import data: {str(e)}")
                return render(request, 'divineImport/import_health_record.html', {'form': form})
    else:
        form = HealthRecordImportForm()

    return render(request, 'divineImport/import_health_record.html', {'form': form})




def import_remote_company_data(request):
    if request.method == 'POST':
        form = RemoteCompanyImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            try:
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active

                # Read headers
                headers = [cell.value.strip() for cell in sheet[1]]  # Trim headers
                required_headers = ['name', 'industry', 'sector', 'headquarters', 'Founded', 'Notes']

                # Validate headers
                if headers[:len(required_headers)] != required_headers:
                    messages.error(request, 'Invalid file format')
                    return render(request, 'divineImport/import_remote_company.html', {'form': form})

                # Read data from rows
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    try:
                        # Ensure each header is mapped to a value, defaulting to an empty string if necessary
                        data = dict(zip(headers, (cell.strip() if isinstance(cell, str) else cell for cell in row)))  # Trim data
                        name = data.get('name')
                        industry = data.get('industry', '')
                        sector = data.get('sector', '')
                        headquarters = data.get('headquarters', '')
                        Founded = data.get('Founded', '')
                        Notes = data.get('Notes', '')

                        if name:  # Only proceed if name is present
                            RemoteCompany.objects.create(
                                name=name.strip(),  # Trim name
                                industry=industry.strip(),
                                sector=sector.strip(),
                                headquarters=headquarters.strip(),
                                Founded=Founded.strip(),
                                Notes=Notes.strip()
                            )
                    except IntegrityError:
                        # Skip duplicate entries and continue
                        continue
                    except Exception as e:
                        messages.error(request, f"Failed to import row data: {str(e)}")
                        continue

                return HttpResponseRedirect(reverse('divine_manage_company'))

            except Exception as e:
                messages.error(request, f"Failed to import data: {str(e)}")
                return render(request, 'divineImport/import_remote_company.html', {'form': form})
    else:
        form = RemoteCompanyImportForm()

    return render(request, 'divineImport/import_remote_company.html', {'form': form})




def import_pathodology_record_data(request):
    if request.method == 'POST':
        form = PathodologyRecordImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            try:
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active

                # Read headers
                headers = [cell.value.strip() for cell in sheet[1]]  # Trim headers
                required_headers = ['name', 'description']

                # Validate headers
                if headers[:len(required_headers)] != required_headers:
                    messages.error(request, 'Invalid file format')
                    return render(request, 'divineImport/import_pathodology_record.html', {'form': form})

                # Read data from rows
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    try:
                        # Ensure each header is mapped to a value, defaulting to an empty string if necessary
                        data = dict(zip(headers, (cell.strip() if isinstance(cell, str) else cell for cell in row)))  # Trim data
                        name = data.get('name')
                        description = data.get('description', '')

                        if name:  # Only proceed if name is present
                            PathodologyRecord.objects.create(
                                name=name.strip(),  # Trim name
                                description=description.strip() if description else ''  # Trim description
                            )
                    except IntegrityError:
                        # Skip duplicate entries and continue
                        continue
                    except Exception as e:
                        messages.error(request, f"Failed to import row data: {str(e)}")
                        continue

                return HttpResponseRedirect(reverse('divine_manage_pathodology'))

            except Exception as e:
                messages.error(request, f"Failed to import data: {str(e)}")
                return render(request, 'divineImport/import_pathodology_record.html', {'form': form})
    else:
        form = PathodologyRecordImportForm()

    return render(request, 'divineImport/import_pathodology_record.html', {'form': form})


def import_remote_service_data(request):
    if request.method == 'POST':
        form = RemoteServiceImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            try:
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active

                # Read headers
                headers = [cell.value.strip() for cell in sheet[1]]  # Trim headers
                required_headers = ['name', 'description', 'category']

                # Validate headers
                if headers[:len(required_headers)] != required_headers:
                    messages.error(request, 'Invalid file format')
                    return render(request, 'divineImport/import_remote_service.html', {'form': form})

                # Read data from rows
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    data = dict(zip(headers, (cell.strip() if isinstance(cell, str) else cell for cell in row)))  # Trim data
                    try:
                        RemoteService.objects.create(
                            name=data['name'].strip(),  # Trim name
                            description=data['description'].strip() if data['description'] else '',  # Trim description
                            category=data['category'].strip() if data['category'] else '',  # Trim category
                        )
                    except IntegrityError:
                        # Skip duplicate entries and continue
                        continue
                    except Exception as e:
                        messages.error(request, f"Failed to import row data: {str(e)}")
                        continue

                return HttpResponseRedirect(reverse('divine_remoteservice_list'))

            except Exception as e:
                messages.error(request, f"Failed to import data: {str(e)}")
                return render(request, 'divineImport/import_remote_service.html', {'form': form})
    else:
        form = RemoteServiceImportForm()

    return render(request, 'divineImport/import_remote_service.html', {'form': form})


def import_country_data(request):
    if request.method == 'POST':
        form = CountryImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            try:
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active

                # Read headers
                headers = [cell.value.strip() for cell in sheet[1]]  # Trim headers
                required_headers = ['name']

                # Validate headers
                if headers[:len(required_headers)] != required_headers:
                    messages.error(request, 'Invalid file format')
                    return render(request, 'divineImport/import_country.html', {'form': form})

                # Read data from rows
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    data = dict(zip(headers, (cell.strip() if isinstance(cell, str) else cell for cell in row)))  # Trim data
                    try:
                        Country.objects.create(
                            name=data['name'].strip()  # Trim name
                        )
                    except IntegrityError:
                        # Skip duplicate entries and continue
                        continue
                    except Exception as e:
                        messages.error(request, f"Failed to import row data: {str(e)}")
                        continue

                return HttpResponseRedirect(reverse('divine_manage_country'))

            except Exception as e:
                messages.error(request, f"Failed to import data: {str(e)}")
                return render(request, 'divineImport/import_country.html', {'form': form})
    else:
        form = CountryImportForm()

    return render(request, 'divineImport/import_country.html', {'form': form})


def import_remote_reagent_data(request):
    if request.method == 'POST':
        form = RemoteReagentForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            try:
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active

                # Read headers
                headers = [cell.value.strip() for cell in sheet[1]]  # Trim headers
                required_headers = ['name', 'supplier', 'quantity', 'expiry_date', 'storage_conditions']

                # Validate headers
                if headers[:len(required_headers)] != required_headers:
                    messages.error(request, 'Invalid file format')
                    return render(request, 'divineImport/import_remote_reagent.html', {'form': form})

                # Read data from rows
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    data = dict(zip(headers, (cell.strip() if isinstance(cell, str) else cell for cell in row)))  # Trim data
                    try:
                        RemoteReagent.objects.create(
                            name=data['name'],
                            supplier=data['supplier'],
                            quantity=data['quantity'],
                            expiry_date=data['expiry_date'],
                            storage_conditions=data['storage_conditions']
                        )
                    except IntegrityError:
                        # Skip duplicate entries and continue
                        continue
                    except Exception as e:
                        messages.error(request, f"Failed to import row data: {str(e)}")
                        continue

                return HttpResponseRedirect(reverse('divine_reagent_list'))

            except Exception as e:
                messages.error(request, f"Failed to import data: {str(e)}")
                return render(request, 'divineImport/import_remote_reagent.html', {'form': form})
    else:
        form = RemoteReagentForm()

    return render(request, 'divineImport/import_remote_reagent.html', {'form': form})


def import_remote_equipment_data(request):
    if request.method == 'POST':
        form = RemoteEquipmentForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            try:
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active

                # Read headers
                headers = [cell.value for cell in sheet[1]]
                required_headers = [
                    'name', 'description', 'serial_number', 'manufacturer',
                    'purchase_date', 'warranty_expiry_date', 'location', 'status'
                ]

                # Validate headers
                if headers[:len(required_headers)] != required_headers:
                    messages.error(request, 'Invalid file format')
                    return render(request, 'divineImport/import_remote_equipment.html', {'form': form})

                # Read data from rows
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    data = dict(zip(headers, row))
                    try:
                        RemoteEquipment.objects.create(
                            name=data['name'].strip(),
                            description=data['description'].strip() if data['description'] else None,
                            serial_number=data['serial_number'].strip(),
                            manufacturer=data['manufacturer'].strip() if data['manufacturer'] else None,
                            purchase_date=data['purchase_date'],
                            warranty_expiry_date=data['warranty_expiry_date'],
                            location=data['location'].strip() if data['location'] else None,
                            status=data['status'].strip() if data['status'] else None
                        )
                    except IntegrityError:
                        # Skip duplicate entries and continue
                        continue
                    except Exception as e:
                        messages.error(request, f"Failed to import row data: {str(e)}")
                        continue

                return HttpResponseRedirect(reverse('divine_remote_equipment_list'))

            except Exception as e:
                messages.error(request, f"Failed to import data: {str(e)}")
                return render(request, 'divineImport/import_remote_equipment.html', {'form': form})
    else:
        form = RemoteEquipmentForm()

    return render(request, 'divineImport/import_remote_equipment.html', {'form': form})


def import_diagnosis_data(request):
    if request.method == 'POST':
        form = DiagnosisImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            try:
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active

                # Read headers
                headers = [cell.value for cell in sheet[1]]
                required_headers = ['diagnosis_name', 'diagnosis_code']

                # Validate headers
                if headers[:len(required_headers)] != required_headers:
                    messages.error(request, "Invalid file format")
                    return render(request, 'divineImport/import_diagnosis.html', {'form': form})

                # Read data from rows
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    data = dict(zip(headers, row))
                    
                    # Trim whitespace from data
                    data = {key: value.strip() if isinstance(value, str) else value for key, value in data.items()}
                    
                    try:
                        Diagnosis.objects.create(
                            diagnosis_name=data['diagnosis_name'],
                            diagnosis_code=data['diagnosis_code']
                        )
                    except IntegrityError:
                        # Skip duplicate entries and continue
                        continue

                return HttpResponseRedirect(reverse('divine_diagnosis_list'))

            except Exception as e:
                messages.error(request, f"Failed to import data: {str(e)}")
                return render(request, 'divineImport/import_diagnosis.html', {'form': form})
    else:
        form = DiagnosisImportForm()

    return render(request, 'divineImport/import_diagnosis.html', {'form': form})




