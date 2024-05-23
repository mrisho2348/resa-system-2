import openpyxl
from django.http import HttpResponse
from clinic.models import AmbulanceRoute, AmbulanceVehicleOrder, Category, Country, Diagnosis, DiseaseRecode, HealthRecord, HospitalVehicle, InsuranceCompany, MedicineUnitMeasure, PathodologyRecord, PrescriptionFrequency, RemoteCompany, RemoteEquipment, RemoteMedicine, RemoteReagent, RemoteService, Service, Staffs, Supplier


def download_disease_recode_excel_template(request):
    # Create a new Workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "DiseaseRecode Template"
    
    # Define column headers from model fields
    excluded_fields = ['created_at', 'updated_at']
    model_fields = [field.name for field in DiseaseRecode._meta.get_fields() 
                    if not field.auto_created and not field.is_relation and field.name not in excluded_fields]

    # Add headers to the first row
    for col_num, column_title in enumerate(model_fields, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=disease_recode_template.xlsx'
    workbook.save(response)

    return response


def download_health_record_excel_template(request):
    # Create a new Workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "HealthRecord Template"
    
    # Define column headers from model fields
    excluded_fields = ['created_at', 'updated_at']
    model_fields = [field.name for field in HealthRecord._meta.get_fields() 
                    if not field.auto_created and not field.is_relation and field.name not in excluded_fields]

    # Add headers to the first row
    for col_num, column_title in enumerate(model_fields, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=health_record_template.xlsx'
    workbook.save(response)

    return response

def download_remote_company_excel_template(request):
    # Create a new Workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "RemoteCompany Template"
    
    # Define column headers from model fields
    excluded_fields = ['created_at', 'updated_at']
    model_fields = [field.name for field in RemoteCompany._meta.get_fields() 
                    if not field.auto_created and not field.is_relation and field.name not in excluded_fields]

    # Add headers to the first row
    for col_num, column_title in enumerate(model_fields, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=remote_company_template.xlsx'
    workbook.save(response)

    return response

def download_pathodology_record_excel_template(request):
    # Create a new Workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "PathodologyRecord Template"
    
    # Define column headers from model fields
    excluded_fields = ['created_at', 'updated_at']
    model_fields = [field.name for field in PathodologyRecord._meta.get_fields() 
                    if not field.auto_created and not field.is_relation and field.name not in excluded_fields]

    # Add headers to the first row
    for col_num, column_title in enumerate(model_fields, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=pathodology_record_template.xlsx'
    workbook.save(response)

    return response

def download_supplier_excel_template(request):
    # Create a new Workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Supplier Template"
    
    # Define column headers from model fields
    excluded_fields = ['created_at', 'updated_at']
    model_fields = [field.name for field in Supplier._meta.get_fields() 
                    if not field.auto_created and not field.is_relation and field.name not in excluded_fields]

    # Add headers to the first row
    for col_num, column_title in enumerate(model_fields, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=supplier_template.xlsx'
    workbook.save(response)

    return response

def download_category_excel_template(request):
    # Create a new Workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Category Template"
    
    # Define column headers from model fields
    excluded_fields = ['created_at', 'updated_at']
    model_fields = [field.name for field in Category._meta.get_fields() 
                    if not field.auto_created and not field.is_relation and field.name not in excluded_fields]

    # Add headers to the first row
    for col_num, column_title in enumerate(model_fields, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=category_template.xlsx'
    workbook.save(response)

    return response


def download_medicine_unit_measure_excel_template(request):
    # Create a new Workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "MedicineUnitMeasure Template"
    
    # Define column headers from model fields
    excluded_fields = ['created_at', 'updated_at']
    model_fields = [field.name for field in MedicineUnitMeasure._meta.get_fields() 
                    if not field.auto_created and not field.is_relation and field.name not in excluded_fields]

    # Add headers to the first row
    for col_num, column_title in enumerate(model_fields, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=medicine_unit_measure_template.xlsx'
    workbook.save(response)

    return response


def download_service_excel_template(request):
    # Create a new Workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Service Template"
    
    # Define column headers from model fields
    excluded_fields = ['created_at', 'updated_at']
    model_fields = [field.name for field in Service._meta.get_fields() 
                    if not field.auto_created and not field.is_relation and field.name not in excluded_fields]

    # Add headers to the first row
    for col_num, column_title in enumerate(model_fields, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=service_template.xlsx'
    workbook.save(response)

    return response

def download_insurance_company_excel_template(request):
    # Create a new Workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "InsuranceCompany Template"
    
    # Define column headers from model fields
    excluded_fields = ['created_at', 'updated_at']
    model_fields = [field.name for field in InsuranceCompany._meta.get_fields() 
                    if not field.auto_created and not field.is_relation and field.name not in excluded_fields]

    # Add headers to the first row
    for col_num, column_title in enumerate(model_fields, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=insurance_company_template.xlsx'
    workbook.save(response)

    return response


def download_staffs_excel_template(request):
    # Create a new Workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Staffs Template"
    
    # Define column headers from model fields
    excluded_fields = ['id', 'admin', 'created_at', 'updated_at']
    model_fields = [field.name for field in Staffs._meta.get_fields() 
                    if not field.auto_created and not field.is_relation and field.name not in excluded_fields]

    # Add headers to the first row
    for col_num, column_title in enumerate(model_fields, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=staffs_template.xlsx'
    workbook.save(response)

    return response

def download_prescription_frequency_excel_template(request):
    # Create a new Workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "PrescriptionFrequency Template"
    
    # Define column headers from model fields
    excluded_fields = ['created_at', 'updated_at']
    model_fields = [field.name for field in PrescriptionFrequency._meta.get_fields() 
                    if not field.auto_created and not field.is_relation and field.name not in excluded_fields]

    # Add headers to the first row
    for col_num, column_title in enumerate(model_fields, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=prescription_frequency_template.xlsx'
    workbook.save(response)

    return response

def download_ambulance_vehicle_order_excel_template(request):
    # Create a new Workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "AmbulanceVehicleOrder Template"
    
    # Define column headers from model fields
    excluded_fields = ['created_at', 'updated_at']
    model_fields = [field.name for field in AmbulanceVehicleOrder._meta.get_fields() 
                    if not field.auto_created and not field.is_relation and field.name not in excluded_fields]

    # Add headers to the first row
    for col_num, column_title in enumerate(model_fields, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=ambulance_vehicle_order_template.xlsx'
    workbook.save(response)

    return response

def download_ambulance_route_excel_template(request):
    # Create a new Workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "AmbulanceRoute Template"
    
    # Define column headers from model fields
    excluded_fields = ['created_at', 'updated_at', 'total']
    model_fields = [field.name for field in AmbulanceRoute._meta.get_fields() 
                    if not field.auto_created and not field.is_relation and field.name not in excluded_fields]

    # Add headers to the first row
    for col_num, column_title in enumerate(model_fields, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=ambulance_route_template.xlsx'
    workbook.save(response)

    return response

def download_hospital_vehicle_excel_template(request):
    # Create a new Workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "HospitalVehicle Template"
    
    # Define column headers from model fields
    excluded_fields = ['created_at', 'updated_at']
    model_fields = [field.name for field in HospitalVehicle._meta.get_fields() 
                    if not field.auto_created and not field.is_relation and field.name not in excluded_fields]

    # Add headers to the first row
    for col_num, column_title in enumerate(model_fields, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=hospital_vehicle_template.xlsx'
    workbook.save(response)

    return response

def download_remote_medicine_excel_template(request):
    # Create a new Workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "RemoteMedicine Template"
    
    # Define column headers from model fields
    excluded_fields = ['created_at', 'updated_at', 'total_buying_price','remain_quantity']
    model_fields = [field.name for field in RemoteMedicine._meta.get_fields() 
                    if not field.auto_created and not field.is_relation and field.name not in excluded_fields]

    # Add headers to the first row
    for col_num, column_title in enumerate(model_fields, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=remote_medicine_template.xlsx'
    workbook.save(response)

    return response

def download_remote_service_excel_template(request):
    # Create a new Workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "RemoteService Template"
    
    # Define column headers from model fields
    excluded_fields = ['created_at', 'updated_at']
    model_fields = [field.name for field in RemoteService._meta.get_fields() 
                    if not field.auto_created and not field.is_relation and field.name not in excluded_fields]

    # Add headers to the first row
    for col_num, column_title in enumerate(model_fields, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=remote_service_template.xlsx'
    workbook.save(response)

    return response

def download_country_excel_template(request):
    # Create a new Workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Country Template"
    
    # Define column headers from model fields
    excluded_fields = ['created_at', 'updated_at']
    model_fields = [field.name for field in Country._meta.get_fields() 
                    if not field.auto_created and not field.is_relation and field.name not in excluded_fields]

    # Add headers to the first row
    for col_num, column_title in enumerate(model_fields, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=country_template.xlsx'
    workbook.save(response)

    return response

def download_remote_reagent_excel_template(request):
    # Create a new Workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Reagent Template"
    
    # Define column headers from model fields
    excluded_fields = ['id', 'date_received']
    model_fields = [field.name for field in RemoteReagent._meta.get_fields()
                    if not field.auto_created and not field.is_relation and field.name not in excluded_fields]

    # Add headers to the first row
    for col_num, column_title in enumerate(model_fields, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=remote_reagent_template.xlsx'
    workbook.save(response)

    return response

def download_remote_equipment_excel_template(request):
    # Create a new Workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Equipment Template"
    
    # Define column headers from model fields
    excluded_fields = ['id', 'date_added']
    model_fields = [field.name for field in RemoteEquipment._meta.get_fields()
                    if not field.auto_created and not field.is_relation and field.name not in excluded_fields]

    # Add headers to the first row
    for col_num, column_title in enumerate(model_fields, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=equipment_template.xlsx'
    workbook.save(response)

    return response

def download_diagnosis_excel_template(request):
    # Create a new Workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Diagnosis Template"
    
    # Define column headers from model fields
    excluded_fields = ['id', 'created_at', 'updated_at']
    model_fields = [field.name for field in Diagnosis._meta.get_fields()
                    if not field.auto_created and not field.is_relation and field.name not in excluded_fields]

    # Add headers to the first row
    for col_num, column_title in enumerate(model_fields, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=diagnosis_template.xlsx'
    workbook.save(response)

    return response