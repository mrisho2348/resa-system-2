# views.py
import logging
from django.shortcuts import redirect, render
from django.contrib import messages
from django.db import IntegrityError
import openpyxl
from clinic.models import AmbulanceActivity, AmbulanceRoute,   MedicineRoute, MedicineUnitMeasure, PrescriptionFrequency,  Diagnosis, DiseaseRecode, Equipment,  Medicine, PathodologyRecord,  Patients,  Procedure, Reagent, Referral, RemoteCompany, Service
from .resources import AmbulanceActivityResource, AmbulanceRouteResource,  CompanyResource,  DiagnosisResource, DiseaseRecodeResource,  EquipmentResource,    MedicineRouteResource, MedicineUnitMeasureResource, PathologyRecordResource,  PatientsResource, PrescriptionFrequencyResource,  ProcedureResource, ReagentResource, ReferralResource, ServiceResource
from .forms import ImportAmbulanceActivityForm, ImportAmbulanceRouteForm,  ImportCompanyForm,  ImportDiagnosisForm, ImportDiseaseForm, ImportEquipmentForm,     ImportMedicineForm, ImportMedicineRouteForm, ImportMedicineUnitMeasureForm, ImportPathologyRecordForm,  ImportPatientsForm,  ImportPrescriptionFrequencyForm, ImportProcedureForm, ImportReagentForm, ImportReferralForm, ImportServiceForm
from tablib import Dataset
from django.contrib.auth.decorators import login_required
logger = logging.getLogger(__name__)

@login_required  
def import_disease_recode(request):
    if request.method == 'POST':
        form = ImportDiseaseForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                disease_recode_resource = DiseaseRecodeResource()
                new_disease_recodes = request.FILES['disease_recode_file']
                
                # Use tablib to load the imported data
                dataset = Dataset()
                imported_data = dataset.load(new_disease_recodes.read(), format='xlsx')  # Adjust format accordingly

                for data in imported_data:
                    disease_recode = DiseaseRecode(
                        disease_name=data[0],
                        code=data[1],
                        # Add other fields accordingly
                    )
                    disease_recode.save()

                return redirect('clinic:manage_disease') 
            except Exception as e:
                messages.error(request, f'An error occurred: {e}')
        else:
            messages.error(request, 'Form is not valid. Please check the file and try again.')
    else:
        form = ImportDiseaseForm()
    return render(request, 'hod_template/import_disease_recode.html', {'form': form})



@login_required  
def import_prescription_frequency(request):
    if request.method == 'POST':
        form = ImportPrescriptionFrequencyForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                resource = PrescriptionFrequencyResource()
                new_record = request.FILES['records_file']
                
                # Use tablib to load the imported data
                dataset = Dataset()
                imported_data = dataset.load(new_record.read(), format='xlsx')  # Assuming you are using xlsx, adjust accordingly

                for data in imported_data:
                     recode = PrescriptionFrequency(
                        name=data[0],
                        interval=data[1],
                        description=data[2],                      
                        # Add other fields accordingly
                    )
                     recode.save()

                return redirect('clinic:prescription_frequency_list') 
            except Exception as e:
                messages.error(request, f'An error occurred: {e}')

    else:
        form = ImportPrescriptionFrequencyForm()
    return render(request, 'hod_template/import_prescription_frequency.html', {'form': form})

@login_required  
def import_ambulance_route_list(request):
    if request.method == 'POST':
        form = ImportAmbulanceRouteForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                resource = AmbulanceRouteResource()
                new_record = request.FILES['records_file']
                
                # Use tablib to load the imported data
                dataset = Dataset()
                imported_data = dataset.load(new_record.read(), format='xlsx')  # Assuming you are using xlsx, adjust accordingly

                for data in imported_data:
                     recode = AmbulanceRoute(
                        from_location=data[0],
                        to_location=data[1],
                        distance=data[2],                      
                        cost=data[3],                      
                        profit=data[4],                      
                        advanced_ambulance_cost=data[5],                     
                    
                    )
                     recode.save()

                return redirect('clinic:ambulance_route_list') 
            except Exception as e:
                messages.error(request, f'An error occurred: {e}')

    else:
        form = ImportAmbulanceRouteForm()
    return render(request, 'hod_template/import_ambulance_route_list.html', {'form': form})

@login_required  
def import_ambulance_activity_list(request):
    if request.method == 'POST':
        form = ImportAmbulanceActivityForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                resource = AmbulanceActivityResource()
                new_record = request.FILES['records_file']                
                # Use tablib to load the imported data
                dataset = Dataset()
                imported_data = dataset.load(new_record.read(), format='xlsx')  # Assuming you are using xlsx, adjust accordingly

                for data in imported_data:
                     recode = AmbulanceActivity(
                        name=data[0],
                        cost=data[1],
                        profit=data[2],                    
                   
                    )
                     recode.save()

                return redirect('clinic:ambulance_activity_list') 
            except Exception as e:
                messages.error(request, f'An error occurred: {e}')

    else:
        form = ImportAmbulanceActivityForm()
    return render(request, 'hod_template/import_ambulance_activity_list.html', {'form': form})



@login_required  
def import_equipment(request):
    if request.method == 'POST':
        form = ImportEquipmentForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                resource = EquipmentResource()
                new_equipment = request.FILES['equipment_records_file']
                
                # Use tablib to load the imported data
                dataset = Dataset()
                imported_data = dataset.load(new_equipment.read(), format='xlsx')  # Assuming you are using xlsx, adjust accordingly

                for data in imported_data:
                     equipment_record = Equipment(
                        name=data[0],
                        description=data[1],
                        manufacturer=data[2],
                        serial_number=data[3],
                        acquisition_date=data[4],
                        warranty_expiry_date=data[5],
                        location=data[6],
                      
                    )
                     equipment_record.save()

                return redirect('clinic:equipment_list') 
            except Exception as e:
                messages.error(request, f'An error occurred: {e}')

    else:
        form = ImportEquipmentForm()

    return render(request, 'hod_template/import_equipment.html', {'form': form})



@login_required  
def import_reagent(request):
    if request.method == 'POST':
        form = ImportReagentForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                resource = ReagentResource()
                new_reagent = request.FILES['reagent_records_file']
                
                # Use tablib to load the imported data
                dataset = Dataset()
                imported_data = dataset.load(new_reagent.read(), format='xlsx')  # Assuming you are using xlsx, adjust accordingly

                for data in imported_data:
                     reagent_record = Reagent(
                        name=data[0],
                        expiration_date=data[1],
                        manufacturer=data[2],
                        lot_number=data[3],
                        storage_conditions=data[4],
                        quantity_in_stock=data[5],
                        price_per_unit=data[6],
                        remaining_quantity=data[5],
                       
                  
                      
                    )
                     reagent_record.save()

                return redirect('clinic:reagent_list') 
            except Exception as e:
                messages.error(request, f'An error occurred: {e}')

    else:
        form = ImportReagentForm()

    return render(request, 'hod_template/import_reagent.html', {'form': form})




@login_required  
def import_companies(request):
    if request.method == 'POST':
        form = ImportCompanyForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                resource = CompanyResource()
                new_companies = request.FILES['company_file']
                
                # Use tablib to load the imported data
                dataset = resource.export()
                imported_data = dataset.load(new_companies.read(), format='xlsx')  # Assuming you are using xlsx, adjust accordingly

                for data in imported_data:
                      company_recode = RemoteCompany(
                        name=data[0],
                        code=data[1],
                        category=data[2],
                      
                        # Add other fields accordingly
                    )
                      company_recode.save()

                return redirect('clinic:manage_company') 
            except Exception as e:
                logger.error(f"Error adding company: {str(e)}")
                messages.error(request, f'An error occurred: {e}')

    else:
        form = ImportCompanyForm()

    return render(request, 'hod_template/import_companies.html', {'form': form})


@login_required  
def import_pathology_records(request):
    if request.method == 'POST':
        form = ImportPathologyRecordForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                resource = PathologyRecordResource()
                new_records = request.FILES['pathology_records_file']
                
                # Use tablib to load the imported data
                dataset = resource.export()
                imported_data = dataset.load(new_records.read(), format='xlsx')  # Assuming you are using xlsx, adjust accordingly

                for data in imported_data:
                     pathodology_recode = PathodologyRecord(
                        name=data[0],
                        description=data[1],                     
                      
                        # Add other fields accordingly
                    )
                     pathodology_recode.save()

                return redirect('clinic:manage_pathodology') 
            except Exception as e:
                messages.error(request, f'An error occurred: {e}')

    else:
        form = ImportPathologyRecordForm()

    return render(request, 'hod_template/import_pathology_records.html', {'form': form})



@login_required  
def import_diagnosis_records(request):
    if request.method == 'POST':
        form = ImportDiagnosisForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                resource = DiagnosisResource()
                new_records = request.FILES['diagnosis_records_file']
                
                # Use tablib to load the imported data
                dataset = resource.export()
                imported_data = dataset.load(new_records.read(), format='xlsx')  # Assuming you are using xlsx, adjust accordingly

                for data in imported_data:
                     diagnosis_record = Diagnosis(
                 
                        diagnosis_name=data[0],                     
                        diagnosis_code=data[1],                     
                                          
                                         
                       )
                     diagnosis_record.save()

                return redirect('clinic:diagnosis_list') 
            except Exception as e:
                messages.error(request, f'An error occurred: {e}')

    else:
        form = ImportDiagnosisForm()

    return render(request, 'hod_template/import_diagnosis.html', {'form': form})

@login_required  
def import_patient_records(request):
    if request.method == 'POST':
        form = ImportPatientsForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                resource = PatientsResource()
                new_records = request.FILES['patient_records_file']

                # Use tablib to load the imported data
                dataset = resource.export()
                imported_data = dataset.load(new_records.read(), format='xlsx')  # Assuming you are using xlsx, adjust accordingly
                
                for data in imported_data:
                    try:
                        patient_record = Patients.objects.create(
                            email=data[1],
                            fullname=data[0],                     
                            dob=data[2],                     
                            gender=data[3],                     
                            phone=data[4],                     
                            address=data[5],                     
                            nationality=data[6],                     
                            company=data[7],                     
                            marital_status=data[8],                     
                            patient_type=data[9],                     
                        )
                    except IntegrityError:
                        messages.warning(request, f'Duplicate entry found for {data[0]}. Skipping this record.')
                        continue

                return redirect('clinic:manage_patient') 
            except Exception as e:
                messages.error(request, f'An error occurred: {e}')

    else:
        form = ImportPatientsForm()

    return render(request, 'hod_template/import_patients.html', {'form': form})

@login_required  
def import_service_records(request):
    if request.method == 'POST':
        form = ImportServiceForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                resource = ServiceResource()
                new_records = request.FILES['service_records_file']

                # Use tablib to load the imported data
                dataset = resource.export()
                imported_data = dataset.load(new_records.read(), format='xlsx')  # Assuming you are using xlsx, adjust accordingly
                
                for data in imported_data:
                    try:
                        service_record = Service.objects.create(                         
                            type_service=data[5],                     
                            name=data[0],                     
                            description=data[4],                     
                            cash_cost=data[2],                     
                            insurance_cost=data[3],                     
                            nhif_cost=data[4],                     
                            coverage=data[1],                     
                                              
                        )
                    except IntegrityError:
                        messages.warning(request, f'Duplicate entry found for {data[0]}. Skipping this record.')
                        continue

                return redirect('clinic:manage_service') 
            except Exception as e:
                messages.error(request, f'An error occurred: {e}')

    else:
        form = ImportServiceForm()

    return render(request, 'hod_template/import_service.html', {'form': form})

@login_required  
def import_medicine_routes_records(request):
    if request.method == 'POST':
        form = ImportMedicineRouteForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                resource = MedicineRouteResource()
                new_records = request.FILES['records_file']

                # Use tablib to load the imported data
                dataset = resource.export()
                imported_data = dataset.load(new_records.read(), format='xlsx')  # Assuming you are using xlsx, adjust accordingly
                
                for data in imported_data:
                    try:
                        new_record = MedicineRoute.objects.create(                        
                            name=data[0],                     
                            explanation=data[1],                    
                           )
                    except IntegrityError:
                        messages.warning(request, f'Duplicate entry found for {data[0]}. Skipping this record.')
                        continue

                return redirect('clinic:medicine_routes') 
            except Exception as e:
                messages.error(request, f'An error occurred: {e}')

    else:
        form = ImportMedicineRouteForm()

    return render(request, 'hod_template/import_medicine_routes.html', {'form': form})

@login_required  
def import_medicine_unit_measures_records(request):
    if request.method == 'POST':
        form = ImportMedicineUnitMeasureForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                resource = MedicineUnitMeasureResource()
                new_records = request.FILES['records_file']

                # Use tablib to load the imported data
                dataset = resource.export()
                imported_data = dataset.load(new_records.read(), format='xlsx')  # Assuming you are using xlsx, adjust accordingly
                
                for data in imported_data:
                    try:
                        new_record = MedicineUnitMeasure.objects.create(                        
                            name=data[0],                     
                            short_name=data[1],                    
                            application_user=data[2],                    
                           )
                    except IntegrityError:
                        messages.warning(request, f'Duplicate entry found for {data[0]}. Skipping this record.')
                        continue

                return redirect('clinic:medicine_routes') 
            except Exception as e:
                messages.error(request, f'An error occurred: {e}')

    else:
        form = ImportMedicineUnitMeasureForm()

    return render(request, 'hod_template/import_medicine_unit_measures.html', {'form': form})


@login_required  
def import_referral_records(request):
    if request.method == 'POST':
        form = ImportReferralForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                resource = ReferralResource()
                new_records = request.FILES['referral_records_file']

                # Use tablib to load the imported data
                dataset = resource.export()
                imported_data = dataset.load(new_records.read(), format='xlsx')  # Assuming you are using xlsx, adjust accordingly
                
                for data in imported_data:
                    try:
                        referral_record =Referral.objects.create(
                            patient=Patients.objects.get(mrn=data[0]),
                            source_location=data[1],                     
                            destination_location=data[2],                     
                            reason=data[3],                     
                            notes=data[4],                     
                                           
                        )
                    except IntegrityError:
                        messages.warning(request, f'Duplicate entry found for {data[0]}. Skipping this record.')
                        continue

                return redirect('clinic:manage_referral') 
            except Exception as e:
                messages.error(request, f'An error occurred: {e}')

    else:
        form = ImportReferralForm()

    return render(request, 'hod_template/import_referral.html', {'form': form})

@login_required  
def import_procedure_records(request):
    if request.method == 'POST':
        form = ImportProcedureForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                resource = ProcedureResource()
                new_records = request.FILES['procedure_records_file']

                # Use tablib to load the imported data
                dataset = resource.export()
                imported_data = dataset.load(new_records.read(), format='xlsx')  # Assuming you are using xlsx, adjust accordingly
                
                for data in imported_data:
                    try:
                        procedure_record = Procedure.objects.create(
                            patient=Patients.objects.get(mrn=data[0]),
                            name=data[1],                     
                            description=data[2],                     
                            duration_time=data[3],                     
                            equipments_used=data[4],                     
                            cost=data[5],                     
                          )
                    except IntegrityError:
                        messages.warning(request, f'Duplicate entry found for {data[0]}. Skipping this record.')
                        continue

                return redirect('clinic:patient_procedure_view') 
            except Exception as e:
                messages.error(request, f'An error occurred: {e}')

    else:
        form = ImportProcedureForm()

    return render(request, 'hod_template/import_procedure.html', {'form': form})

@login_required
def import_medicine_records(request):
    if request.method == 'POST':
        form = ImportMedicineForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['medicine_records_file']
            try:
                # Check if the uploaded file is an Excel file
                if not excel_file.name.endswith('.xlsx'):
                    messages.error(request, 'Please upload a valid Excel file.')
                    return render(request, 'hod_template/import_medicine_records.html', {'form': form})

                # Load the Excel workbook
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active

                # Iterate through rows and create or update Medicine objects
                for row_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if row_num == 1:  # Skip header row
                        continue
                    # Check if a Medicine object with the same batch number already exists
                    batch_number = row[6]
                    if Medicine.objects.filter(batch_number=batch_number).exists():
                        messages.warning(request, f'Skipping drug with batch number {batch_number} as it already exists.')
                        continue

                    # Create Medicine object from Excel data
                    # Handle None values in the Excel file
                    cash_cost = float(row[8]) if row[8] is not None else 0.0
                    nhif_cost = float(row[10]) if row[10] is not None else 0.0
                    buying_price = float(row[11]) if row[11] is not None else 0.0

                    medicine = Medicine(
                        drug_name=row[0],
                        drug_type=row[1],
                        formulation_unit=row[2],
                        manufacturer=row[3],
                        quantity=int(row[4]),
                        remain_quantity=int(row[4]),
                        dividable=row[5],
                        batch_number=batch_number,
                        expiration_date=row[7],
                        cash_cost=cash_cost,
                        insurance_cost=0.0,  # Set default value
                        nhif_cost=nhif_cost,
                        buying_price=buying_price,
                    )
                    medicine.save()

                messages.success(request, 'Medicine data imported successfully.')
                return redirect('clinic:medicine_list')

            except Exception as e:
                # Handle any exceptions that may occur during file processing
                messages.error(request, f'An error occurred: {e}')
        else:
            messages.error(request, 'Please upload a valid Excel file.')
    else:
        form = ImportMedicineForm()

    return render(request, 'hod_template/import_medicine_records.html', {'form': form})


def generate_mrn():
    # Retrieve the last patient's MRN from the database
    last_patient = Patients.objects.last()

    # Extract the numeric part from the last MRN, or start from 0 if there are no patients yet
    last_mrn_number = int(last_patient.mrn.split('-')[-1]) if last_patient else 0

    # Increment the numeric part for the new patient
    new_mrn_number = last_mrn_number + 1

    # Format the MRN with leading zeros and concatenate with the prefix "PAT-"
    new_mrn = f"PAT-{new_mrn_number:05d}"

    return new_mrn

def import_service_data(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            # Load the workbook
            workbook = openpyxl.load_workbook(excel_file)
            # Get the active worksheet
            sheet = workbook.active

            # Iterate through each row in the worksheet starting from the second row (skipping headers)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Extract data from the row
                name, type_service, coverage, description, cash_cost, insurance_cost, nhif_cost, department = row

                # Check if a similar record already exists
                existing_service = Service.objects.filter(
                    name=name,
                    type_service=type_service,
                    coverage=coverage,
                    description=description,
                    cash_cost=cash_cost,
                    insurance_cost=insurance_cost,
                    nhif_cost=nhif_cost,
                    department=department
                ).first()

                # If a similar record does not exist, create a new Service object
                if not existing_service:
                    service = Service.objects.create(
                        name=name,
                        type_service=type_service,
                        coverage=coverage,
                        description=description,
                        cash_cost=cash_cost,
                        insurance_cost=insurance_cost,
                        nhif_cost=nhif_cost,
                        department=department
                    )

                    # Save the Service object
                    service.save()

            return redirect('clinic:manage_service') 
        except Exception as e:
            messages.error(request, f'An error occurred: {e}')
            return render(request, 'hod_template/import_service.html')
    else:
        return render(request, 'hod_template/import_service.html')

